"""
pmo_parser.py — Parser robusto de PDFs do PMO/ONS → Excel de validação
=======================================================================
Suporta qualquer nomenclatura de arquivo — extrai período, tipo e revisão
diretamente do CONTEÚDO do PDF, não do nome do arquivo.

Tipos suportados:
  - Relatório Executivo PMO (estrutura completa: CMO, ENA, EAR, carga, térmico)
  - Sumário Executivo PMO  (estrutura reduzida: CMO, ENA, EAR)
  - PDFs de 2020/2021 com estruturas ligeiramente diferentes

Uso:
    python pmo_parser.py
    python pmo_parser.py --dir data/ons/PMOs --out data/ons/PMOs/validacao_pmo.xlsx
"""
from __future__ import annotations
import re, sys, argparse
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PMO_DIR    = Path("data/ons/PMOs")
OUTPUT_XLS = PMO_DIR / "validacao_pmo.xlsx"

# Meses BR → número
MESES = {"janeiro":1,"fevereiro":2,"marco":3,"março":3,"abril":4,"maio":5,
          "junho":6,"julho":7,"agosto":8,"setembro":9,"outubro":10,
          "novembro":11,"dezembro":12,
          "jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,
          "jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12}

CG="C8A44D"; CD="0B0F14"; CP="111827"; CP2="0D1B2A"
CW="F3F4F6"; CRE="7F1D1D"; COR="7C2D12"; CGR="4A3A5F"


# ══════════════════════════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════════════════════════

def _n(s: str) -> Optional[float]:
    """'62.624' → 62624.0  |  '378,67' → 378.67  |  qualquer separador BR"""
    if not s: return None
    s = s.strip()
    # ponto de milhar + vírgula decimal
    if re.match(r"^\d{1,3}(\.\d{3})+(,\d+)?$", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:    return float(s)
    except: return None


def _detect_content(full: str) -> Dict[str, Any]:
    """
    Extrai metadados do CONTEÚDO do PDF — independente do nome do arquivo.
    Retorna: tipo, semana_inicio, semana_fim, mes_num, ano, revisao.
    """
    meta: Dict[str, Any] = {}

    # ── Tipo de documento ────────────────────────────────────────────────────
    if re.search(r"Relatório Executivo|RELATORIO EXECUTIVO", full[:800], re.I):
        meta["tipo_doc"] = "Relatório Executivo"
    elif re.search(r"Sumário Executivo|SUMARIO EXECUTIVO|Sumario Executivo", full[:800], re.I):
        meta["tipo_doc"] = "Sumário Executivo"
    else:
        meta["tipo_doc"] = "Desconhecido"

    # ── Mês e ano do PMO ─────────────────────────────────────────────────────
    # Padrão: "PMO MARÇO 2026" ou "PMO de Março/2026"
    for pat in [
        r"PMO\s+(?:de\s+)?(\w+)[/\s]+(\d{4})",
        r"PMO\s+(\w+)\s+(\d{4})",
        r"Programa Mensal.*?(\w+)[/\s]+(\d{4})",
    ]:
        m = re.search(pat, full[:600], re.I)
        if m:
            mes_str = m.group(1).lower().strip()
            ano_str = m.group(2)
            mes_num = MESES.get(mes_str)
            if mes_num:
                meta["mes_pmo"]  = m.group(1).capitalize()
                meta["mes_num"]  = mes_num
                meta["ano_pmo"]  = int(ano_str)
                break

    # ── Semana operativa ─────────────────────────────────────────────────────
    for pat in [
        # "SEMANA OPERATIVA DE 14/03 A 20/03/2026"
        r"SEMANA OPERATIVA DE\s*(\d{1,2}/\d{2})(?:/\d{2,4})?\s*[Aa]\s*(\d{1,2}/\d{2}/(\d{4}))",
        # "semana de 14/03 a 20/03/2026"
        r"semana.*?(\d{1,2}/\d{2})(?:/\d{2,4})?\s*[Aa]\s*(\d{1,2}/\d{2}/(\d{4}))",
        # "14/03/2026 a 20/03/2026"
        r"(\d{2}/\d{2}/(\d{4}))\s*[Aa]\s*(\d{2}/\d{2}/\d{4})",
    ]:
        m = re.search(pat, full[:1000], re.I)
        if m:
            try:
                ano = meta.get("ano_pmo") or int(m.group(3))
                g = m.groups()
                d1 = g[0].split("/"); d2 = g[1].split("/")
                # normalizar para dd/mm/aaaa
                meta["semana_inicio"] = datetime(ano, int(d1[1]), int(d1[0])).date()
                meta["semana_fim"]    = datetime(int(g[2]), int(d2[1]), int(d2[0])).date()
                break
            except Exception:
                pass

    # ── Revisão ──────────────────────────────────────────────────────────────
    for pat in [
        r"Revis[aã]o\s*(\d+)\s*do PMO",
        r"(?:RV|Rev\.?)\s*(\d+)",
        r"(\d+)[aª]\s*Revis[aã]o",
    ]:
        m = re.search(pat, full[:1500], re.I)
        if m:
            meta["revisao"] = int(m.group(1))
            break

    # ── Número de semana operativa no mês ────────────────────────────────────
    m = re.search(r"(\d+)[aª]\s*[Ss]emana", full[:500], re.I)
    if m: meta["semana_no_mes"] = int(m.group(1))

    return meta


# ══════════════════════════════════════════════════════════════════════════════
# EXTRATORES ROBUSTOS (tolerantes a variações de layout)
# ══════════════════════════════════════════════════════════════════════════════

def _cmo_apres(t: str) -> Dict[str, float]:
    """CMO semana anterior e atual — Apresentação."""
    r = {}
    for sub, ka, kn in [
        ("SE/CO",    "cmo_ant_seco", "cmo_now_seco"),
        ("Sul",      "cmo_ant_s",    "cmo_now_s"),
        ("Nordeste", "cmo_ant_ne",   "cmo_now_ne"),
        ("Norte",    "cmo_ant_n",    "cmo_now_n"),
    ]:
        # "SE/CO: de R$ 395,68/MWh para R$ 378,67/MWh"
        m = re.search(
            rf"{sub}.*?de R\$\s*([\d.,]+)/MWh\s*para R\$\s*([\d.,]+)/MWh",
            t, re.I)
        if m:
            r[ka] = _n(m.group(1)); r[kn] = _n(m.group(2))
    return r


def _cmo_tab7(t: str) -> Dict[str, float]:
    """CMO Tab7: pesada / média patamar / leve / média semanal."""
    r = {}
    for label, prefix in [
        (r"Pesada",        "cmo_pesada"),
        (r"Média\s+",      "cmo_media"),
        (r"Leve",          "cmo_leve"),
        (r"Média Semanal", "cmo_med"),
    ]:
        m = re.search(
            rf"{label}\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)",
            t)
        if m:
            for i, sub in enumerate(["seco","s","ne","n"], 1):
                r[f"{prefix}_{sub}"] = _n(m.group(i))
    return r


def _cmo_semanas(t: str) -> Dict[str, float]:
    """CMO Sem1..5 — série semanal da Figura 17."""
    r = {}
    for sub, key in [("Sudeste","seco"),("Sul","s"),("Nordeste","ne"),("Norte","n")]:
        m = re.search(
            rf"{sub}\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)",
            t, re.I)
        if m:
            for i in range(1, 6):
                r[f"cmo_sem{i}_{key}"] = _n(m.group(i))
    return r


def _ena_tabelas(t: str) -> Dict[str, float]:
    """
    ENA Tab1 (verificada/estimada) e Tab2 (prevista).
    Estratégia: capturar AMBAS as ocorrências do bloco 4×4 e identificar
    qual é qual pelo contexto (ENAs previstas = Tab2).
    """
    r = {}
    pat4x4 = (
        r"SE/CO\s+([\d.,]+)\s+(\d+)\s+([\d.,]+)\s+(\d+)\n"
        r"S\s+([\d.,]+)\s+(\d+)\s+([\d.,]+)\s+(\d+)\n"
        r"NE\s+([\d.,]+)\s+(\d+)\s+([\d.,]+)\s+(\d+)\n"
        r"N\s+([\d.,]+)\s+(\d+)\s+([\d.,]+)\s+(\d+)"
    )
    all_m = list(re.finditer(pat4x4, t))

    def _fill(m, prefix_sem, prefix_mes):
        g = m.groups()
        for i, sub in enumerate(["seco","s","ne","n"]):
            b = i * 4
            r[f"{prefix_sem}_{sub}_mw"]  = _n(g[b])
            r[f"{prefix_sem}_{sub}_mlt"] = _n(g[b+1])
            r[f"{prefix_mes}_{sub}_mw"]  = _n(g[b+2])
            r[f"{prefix_mes}_{sub}_mlt"] = _n(g[b+3])

    # Identificar qual ocorrência é Tab1 e qual é Tab2
    for m in all_m:
        ctx_before = t[max(0, m.start()-300):m.start()]
        if re.search(r"ENAs? previstas?|Tabela 2|Previs[aã]o.*?ENA", ctx_before, re.I):
            _fill(m, "ena_prev_sem", "ena_prev_mes")
        elif re.search(r"Tend[eê]ncia|Tabela 1|verificad|estimad", ctx_before, re.I):
            _fill(m, "ena_verif_ant", "ena_estim_cur")
        else:
            # fallback: 1a ocorrência = Tab1 (verificada), 2a = Tab2 (prevista)
            if "ena_verif_ant_seco_mw" not in r:
                _fill(m, "ena_verif_ant", "ena_estim_cur")
            elif "ena_prev_sem_seco_mw" not in r:
                _fill(m, "ena_prev_sem", "ena_prev_mes")
    return r


def _ear_tab6(t: str) -> Dict[str, float]:
    """EAR inicial — Tabela 6 (ancorando ao contexto certo)."""
    r = {}
    # Ancorar: buscar bloco que contém "%EARm" para evitar Tab3 (MLT em MWmed)
    blk_m = re.search(r"(?:Tabela 6|Armazenamentos? [Ii]niciais?).*?%EARm", t, re.I|re.DOTALL)
    blk = t[blk_m.start():blk_m.start()+600] if blk_m else ""
    if not blk:
        # fallback: buscar padrão com valores <100 (% não MWmed)
        blk = t

    m = re.search(
        r"SE/CO\s+([\d,]+)\s+([\d,]+)\n"
        r"S\s+([\d,]+)\s+([\d,]+)\n"
        r"NE\s+([\d,]+)\s+([\d,]+)\n"
        r"N\s+([\d,]+)\s+([\d,]+)",
        blk)
    if m:
        g = m.groups()
        for i, sub in enumerate(["seco","s","ne","n"]):
            r[f"ear_init_rv1_{sub}"] = _n(g[i*2])
            r[f"ear_init_rv2_{sub}"] = _n(g[i*2+1])
    return r


def _ear_resumo(t: str) -> Dict[str, float]:
    """
    EAR projetada Inic+Sem1..5+FimMes+VEabr — Figuras 19-22.
    Estratégia posicional: [0]=SE/CO, [1]=NE, [2]=Norte (confirmado empiricamente).
    Sul: gráfico rotacionado → usa Tab11.
    """
    r = {}
    pat_ear = (r"EAR\(%EARmax\)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+"
               r"([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)")
    pat_ena = (r"ENA\(%mlt\)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+"
               r"([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)")
    pat_cmo = (r"CMO \(R\$/MWh\)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+"
               r"([\d,]+)\s+([\d,]+)\s+([\d,]+)\s+([\d,]+)")
    cols_ear = ["inic","sem1","sem2","sem3","sem4","sem5","fim_mes","ve_abr"]
    cols7    = ["sem1","sem2","sem3","sem4","sem5","fim_mes","ve_abr"]

    all_ear = list(re.finditer(pat_ear, t, re.I))
    all_ena = list(re.finditer(pat_ena, t, re.I))
    all_cmo = list(re.finditer(pat_cmo, t, re.I))

    for key, idx in [("seco",0), ("ne",1), ("n",2)]:
        if idx < len(all_ear):
            for ci, col in enumerate(cols_ear, 1):
                r[f"ear_{col}_{key}"] = _n(all_ear[idx].group(ci))
        if idx < len(all_ena):
            for ci, col in enumerate(cols7, 1):
                r[f"ena_resumo_{col}_{key}_mlt"] = _n(all_ena[idx].group(ci))
        if idx < len(all_cmo):
            for ci, col in enumerate(cols7, 1):
                r[f"cmo_resumo_{col}_{key}"] = _n(all_cmo[idx].group(ci))

    # Tab11: EAR início e fim por subsistema (Sul + todos)
    m11 = re.search(
        r"SE/CO\s+([\d,]+)\s+([\d,]+)\s*\n"
        r"Sul\s+([\d,]+)\s+([\d,]+)\s*\n"
        r"Nordeste\s+([\d,]+)\s+([\d,]+)\s*\n"
        r"Norte\s+([\d,]+)\s+([\d,]+)",
        t, re.I)
    if m11:
        g = m11.groups()
        for i, sub in enumerate(["seco","s","ne","n"]):
            r[f"ear_init_tab11_{sub}"] = _n(g[i*2])
            r[f"ear_fim_tab11_{sub}"]  = _n(g[i*2+1])
        # Sul projetada usa Tab11 como melhor fonte disponível
        if "ear_inic_s" not in r:
            r["ear_inic_s"]    = _n(g[2])
            r["ear_fim_mes_s"] = _n(g[3])
    return r


def _carga(t: str) -> Dict[str, float]:
    """Carga SIN Tab5 e previsão textual."""
    r = {}
    # Mensal SIN do texto
    m = re.search(r"SIN.*?carga de\s*([\d.]+)\s*MW", t, re.I)
    if m: r["carga_mensal_sin"] = _n(m.group(1))
    # Próxima semana
    m = re.search(r"atingindo\s*\n?([\d.]+)\s*MW m", t, re.I)
    if m: r["carga_prox_sem_sin"] = _n(m.group(1))
    # Tab5: SIN com 5 semanas + mensal + var%
    m = re.search(
        r"SIN\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)"
        r"\s+([\d.,]+)\s+(-?[\d.,]+)%", t)
    if m:
        for i in range(1, 6): r[f"carga_sem{i}_sin"] = _n(m.group(i))
        r["carga_mes_sin_tab5"] = _n(m.group(6))
        r["carga_var_pct_sin"]  = _n(m.group(7))
    # Subsistemas
    for sub, key in [("SE/CO","seco"),("Sul","s"),("Nordeste","ne"),("Norte","n")]:
        m = re.search(
            rf"{sub}\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+\s+[\d.,]+"
            rf"\s+([\d.,]+)\s+(-?[\d.,]+)%", t)
        if m:
            r[f"carga_mes_{key}"]     = _n(m.group(1))
            r[f"carga_var_{key}_pct"] = _n(m.group(2))
    return r


def _termico(t: str) -> Dict[str, float]:
    """Geração térmica por subsistema — totais de despacho."""
    r = {}
    for sub, key in [("SE/CO","seco"),("SUL","s"),("NE","ne"),("NORTE","n")]:
        m = re.search(
            rf"TOTAL {sub}.*?\(\d+\)\s+"
            rf"([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+"
            rf"([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+"
            rf"([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)",
            t, re.I)
        if m:
            r[f"term_inflex_p_{key}"] = _n(m.group(1))
            r[f"term_inflex_m_{key}"] = _n(m.group(2))
            r[f"term_merito_p_{key}"] = _n(m.group(4))
            r[f"term_merito_m_{key}"] = _n(m.group(5))
            r[f"term_total_p_{key}"]  = _n(m.group(7))
            r[f"term_total_m_{key}"]  = _n(m.group(8))
    totais = [r.get(f"term_total_m_{k}") for k in ["seco","s","ne","n"]]
    if all(v is not None for v in totais):
        r["term_total_sin"] = sum(totais)
    return r


def _politica(t: str) -> Dict[str, str]:
    """Política operativa — texto por região."""
    r = {}
    for key, start, end in [
        ("seco", r"Regi[aã]o SE/CO",   r"Regi[aã]o Sul"),
        ("s",    r"Regi[aã]o Sul",      r"Regi[aã]o NE"),
        ("ne",   r"Regi[aã]o NE",       r"Regi[aã]o Norte"),
        ("n",    r"Regi[aã]o Norte",     r"(?:4\.2\.|5\.|6\.)"),
    ]:
        m = re.search(rf"{start}[:\s]*\n[-–\s]*(.*?)(?={end})", t, re.I|re.DOTALL)
        if m:
            r[f"politica_{key}"] = re.sub(r"\s+", " ", m.group(1).strip())[:500]
    return r


# ══════════════════════════════════════════════════════════════════════════════
# PARSER PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def parse_pmo(pdf_path: Path) -> Dict[str, Any]:
    """Extrai todos os campos de um PDF de PMO."""
    print(f"  ► {pdf_path.name}")
    rec: Dict[str, Any] = {"arquivo": pdf_path.name}

    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            full = "\n".join((p.extract_text() or "") for p in pdf.pages)
            rec["n_paginas"] = len(pdf.pages)
    except Exception as e:
        print(f"    ERRO: {e}")
        rec["erro"] = str(e)
        return rec

    # 1. Metadados do conteúdo (sem depender do nome do arquivo)
    meta = _detect_content(full)
    rec.update(meta)

    # 2. Campos econômicos — com graceful fallback por tipo de documento
    rec.update(_cmo_apres(full))
    rec.update(_cmo_tab7(full))
    rec.update(_cmo_semanas(full))
    rec.update(_ena_tabelas(full))
    rec.update(_ear_tab6(full))
    rec.update(_ear_resumo(full))

    # Campos presentes apenas no Relatório Executivo completo
    if rec.get("tipo_doc") != "Sumário Executivo":
        rec.update(_carga(full))
        rec.update(_termico(full))
        rec.update(_politica(full))

    # 3. Diagnóstico de qualidade
    campos_chave = [
        "cmo_med_seco", "cmo_med_s",
        "ena_prev_sem_seco_mw", "ena_prev_sem_seco_mlt",
        "ear_init_rv2_seco",
    ]
    preenchidos = sum(1 for k in campos_chave if rec.get(k) not in (None, ""))
    rec["qualidade_pct"] = round(preenchidos / len(campos_chave) * 100)

    print(f"    tipo={rec.get('tipo_doc','?')} | "
          f"período={rec.get('semana_inicio','?')} → {rec.get('semana_fim','?')} | "
          f"campos={sum(1 for v in rec.values() if v not in (None,''))} | "
          f"qualidade={rec['qualidade_pct']}%")
    return rec


def parse_all(pmo_dir: Path) -> List[Dict]:
    """Lê todos os PDFs com PMO no nome, de qualquer formato."""
    # Padrão amplo: qualquer PDF que contenha "PMO" ou "pmo" no nome
    pdfs = sorted([
        p for p in pmo_dir.glob("*.pdf")
        if re.search(r"pmo|programa.mensal|sumario.*pmo", p.name, re.I)
    ])
    if not pdfs:
        # Fallback: todos os PDFs
        pdfs = sorted(pmo_dir.glob("*.pdf"))

    if not pdfs:
        print(f"⚠  Nenhum PDF encontrado em {pmo_dir}")
        return []

    print(f"Encontrados {len(pdfs)} PDF(s) em {pmo_dir}\n")
    records = [parse_pmo(p) for p in pdfs]

    # Remover 2020 se não desejado (flag configurável)
    return records


# ══════════════════════════════════════════════════════════════════════════════
# DEFINIÇÃO DE CAMPOS E EXCEL
# ══════════════════════════════════════════════════════════════════════════════

CAMPOS = [
    # ── Identificação (do conteúdo, não do nome) ──────────────────────────────
    ("ID","arquivo",              "Arquivo",              "",       CG),
    ("ID","tipo_doc",             "Tipo doc",             "",       CG),
    ("ID","semana_inicio",        "Semana início",        "",       CG),
    ("ID","semana_fim",           "Semana fim",           "",       CG),
    ("ID","mes_pmo",              "Mês PMO",              "",       CG),
    ("ID","ano_pmo",              "Ano PMO",              "",       CG),
    ("ID","revisao",              "Revisão",              "",       CG),
    ("ID","semana_no_mes",        "Sem. no mês",          "",       CG),
    ("ID","n_paginas",            "Páginas",              "",       CG),
    ("ID","qualidade_pct",        "Qualidade %",          "%",      CG),
    # ── CMO apresentação ──────────────────────────────────────────────────────
    ("CMO Ant","cmo_ant_seco",    "SE/CO ant",            "R$/MWh", "1B3A5F"),
    ("CMO Ant","cmo_ant_s",       "Sul ant",              "R$/MWh", "1B3A5F"),
    ("CMO Ant","cmo_ant_ne",      "NE ant",               "R$/MWh", "1B3A5F"),
    ("CMO Ant","cmo_ant_n",       "Norte ant",            "R$/MWh", "1B3A5F"),
    ("CMO Atual","cmo_now_seco",  "SE/CO atual",          "R$/MWh", "1B4A2F"),
    ("CMO Atual","cmo_now_s",     "Sul atual",            "R$/MWh", "1B4A2F"),
    ("CMO Atual","cmo_now_ne",    "NE atual",             "R$/MWh", "1B4A2F"),
    ("CMO Atual","cmo_now_n",     "Norte atual",          "R$/MWh", "1B4A2F"),
    # ── CMO Tab7 ──────────────────────────────────────────────────────────────
    ("CMO Tab7","cmo_pesada_seco","SE/CO Pesada",         "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_pesada_s",   "Sul Pesada",           "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_pesada_ne",  "NE Pesada",            "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_pesada_n",   "Norte Pesada",         "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_media_seco", "SE/CO Média",          "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_media_s",    "Sul Média",            "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_media_ne",   "NE Média",             "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_media_n",    "Norte Média",          "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_leve_seco",  "SE/CO Leve",           "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_leve_s",     "Sul Leve",             "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_leve_ne",    "NE Leve",              "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_leve_n",     "Norte Leve",           "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_med_seco",   "SE/CO Med.Sem.",       "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_med_s",      "Sul Med.Sem.",         "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_med_ne",     "NE Med.Sem.",          "R$/MWh", "163A5F"),
    ("CMO Tab7","cmo_med_n",      "Norte Med.Sem.",       "R$/MWh", "163A5F"),
    # ── CMO Semanas ───────────────────────────────────────────────────────────
    ("CMO Sems","cmo_sem1_seco",  "SE/CO Sem1",           "R$/MWh", "0F2A4F"),
    ("CMO Sems","cmo_sem2_seco",  "SE/CO Sem2",           "R$/MWh", "0F2A4F"),
    ("CMO Sems","cmo_sem3_seco",  "SE/CO Sem3",           "R$/MWh", "0F2A4F"),
    ("CMO Sems","cmo_sem4_seco",  "SE/CO Sem4",           "R$/MWh", "0F2A4F"),
    ("CMO Sems","cmo_sem5_seco",  "SE/CO Sem5",           "R$/MWh", "0F2A4F"),
    ("CMO Sems","cmo_sem1_s",     "Sul Sem1",             "R$/MWh", "0F2A4F"),
    ("CMO Sems","cmo_sem1_ne",    "NE Sem1",              "R$/MWh", "0F2A4F"),
    ("CMO Sems","cmo_sem1_n",     "Norte Sem1",           "R$/MWh", "0F2A4F"),
    # ── ENA verificada ────────────────────────────────────────────────────────
    ("ENA Verif","ena_verif_ant_seco_mw",  "SE/CO ant MW","MWmed",  "0C2A4F"),
    ("ENA Verif","ena_verif_ant_seco_mlt", "SE/CO ant %", "%MLT",   "0C2A4F"),
    ("ENA Verif","ena_estim_cur_seco_mw",  "SE/CO cur MW","MWmed",  "0C2A4F"),
    ("ENA Verif","ena_estim_cur_seco_mlt", "SE/CO cur %", "%MLT",   "0C2A4F"),
    ("ENA Verif","ena_verif_ant_s_mw",     "Sul ant MW",  "MWmed",  "0C2A4F"),
    ("ENA Verif","ena_verif_ant_s_mlt",    "Sul ant %",   "%MLT",   "0C2A4F"),
    ("ENA Verif","ena_verif_ant_ne_mw",    "NE ant MW",   "MWmed",  "0C2A4F"),
    ("ENA Verif","ena_verif_ant_n_mw",     "Norte ant MW","MWmed",  "0C2A4F"),
    # ── ENA prevista ──────────────────────────────────────────────────────────
    ("ENA Prev","ena_prev_sem_seco_mw",    "SE/CO sem MW","MWmed",  "0D3B5F"),
    ("ENA Prev","ena_prev_sem_seco_mlt",   "SE/CO sem %", "%MLT",   "0D3B5F"),
    ("ENA Prev","ena_prev_mes_seco_mlt",   "SE/CO mês %", "%MLT",   "0D3B5F"),
    ("ENA Prev","ena_prev_sem_s_mw",       "Sul sem MW",  "MWmed",  "0D3B5F"),
    ("ENA Prev","ena_prev_sem_s_mlt",      "Sul sem %",   "%MLT",   "0D3B5F"),
    ("ENA Prev","ena_prev_mes_s_mlt",      "Sul mês %",   "%MLT",   "0D3B5F"),
    ("ENA Prev","ena_prev_sem_ne_mw",      "NE sem MW",   "MWmed",  "0D3B5F"),
    ("ENA Prev","ena_prev_sem_ne_mlt",     "NE sem %",    "%MLT",   "0D3B5F"),
    ("ENA Prev","ena_prev_sem_n_mw",       "Norte sem MW","MWmed",  "0D3B5F"),
    ("ENA Prev","ena_prev_sem_n_mlt",      "Norte sem %", "%MLT",   "0D3B5F"),
    # ── EAR Tab6 ──────────────────────────────────────────────────────────────
    ("EAR Init","ear_init_rv2_seco",  "SE/CO init",    "%EARmax","1E4A1F"),
    ("EAR Init","ear_init_rv2_s",     "Sul init",      "%EARmax","1E4A1F"),
    ("EAR Init","ear_init_rv2_ne",    "NE init",       "%EARmax","1E4A1F"),
    ("EAR Init","ear_init_rv2_n",     "Norte init",    "%EARmax","1E4A1F"),
    # ── EAR projetada ─────────────────────────────────────────────────────────
    ("EAR Proj","ear_inic_seco",      "SE/CO Inic",    "%EARmax","2D4A2F"),
    ("EAR Proj","ear_sem1_seco",      "SE/CO Sem1",    "%EARmax","2D4A2F"),
    ("EAR Proj","ear_sem3_seco",      "SE/CO Sem3",    "%EARmax","2D4A2F"),
    ("EAR Proj","ear_fim_mes_seco",   "SE/CO Fim",     "%EARmax","2D4A2F"),
    ("EAR Proj","ear_ve_abr_seco",    "SE/CO VE[+1]",  "%EARmax","2D4A2F"),
    ("EAR Proj","ear_inic_ne",        "NE Inic",       "%EARmax","2D4A2F"),
    ("EAR Proj","ear_fim_mes_ne",     "NE Fim",        "%EARmax","2D4A2F"),
    ("EAR Proj","ear_ve_abr_ne",      "NE VE[+1]",     "%EARmax","2D4A2F"),
    ("EAR Proj","ear_inic_n",         "Norte Inic",    "%EARmax","2D4A2F"),
    ("EAR Proj","ear_fim_mes_n",      "Norte Fim",     "%EARmax","2D4A2F"),
    ("EAR Proj","ear_ve_abr_n",       "Norte VE[+1]",  "%EARmax","2D4A2F"),
    # ── EAR Sul (Tab11) ───────────────────────────────────────────────────────
    ("EAR Sul","ear_init_rv2_s",      "Sul init(T6)",  "%EARmax","1A3A1F"),
    ("EAR Sul","ear_inic_s",          "Sul inic",      "%EARmax","1A3A1F"),
    ("EAR Sul","ear_fim_mes_s",       "Sul fim",       "%EARmax","1A3A1F"),
    ("EAR Sul","ear_fim_tab11_s",     "Sul fim(T11)",  "%EARmax","1A3A1F"),
    # ── Carga ─────────────────────────────────────────────────────────────────
    ("Carga","carga_mensal_sin",      "Mensal SIN",    "MWmed",  "4A3A1E"),
    ("Carga","carga_prox_sem_sin",    "Próx sem SIN",  "MWmed",  "4A3A1E"),
    ("Carga","carga_sem1_sin",        "Sem1 SIN",      "MWmed",  "4A3A1E"),
    ("Carga","carga_sem2_sin",        "Sem2 SIN",      "MWmed",  "4A3A1E"),
    ("Carga","carga_sem3_sin",        "Sem3 SIN",      "MWmed",  "4A3A1E"),
    ("Carga","carga_sem4_sin",        "Sem4 SIN",      "MWmed",  "4A3A1E"),
    ("Carga","carga_sem5_sin",        "Sem5 SIN",      "MWmed",  "4A3A1E"),
    ("Carga","carga_var_pct_sin",     "Var% SIN",      "%",      "4A3A1E"),
    ("Carga","carga_mes_seco",        "Mensal SE/CO",  "MWmed",  "4A3A1E"),
    ("Carga","carga_mes_s",           "Mensal Sul",    "MWmed",  "4A3A1E"),
    ("Carga","carga_mes_ne",          "Mensal NE",     "MWmed",  "4A3A1E"),
    ("Carga","carga_mes_n",           "Mensal Norte",  "MWmed",  "4A3A1E"),
    # ── Geração térmica ───────────────────────────────────────────────────────
    ("Térmico","term_total_sin",      "Total SIN",     "MWmed",  "4A2A1E"),
    ("Térmico","term_total_m_seco",   "SE/CO Total M", "MWmed",  "4A2A1E"),
    ("Térmico","term_total_m_s",      "Sul Total M",   "MWmed",  "4A2A1E"),
    ("Térmico","term_total_m_ne",     "NE Total M",    "MWmed",  "4A2A1E"),
    ("Térmico","term_total_m_n",      "Norte Total M", "MWmed",  "4A2A1E"),
    ("Térmico","term_merito_m_seco",  "SE/CO Mérito M","MWmed",  "4A2A1E"),
    ("Térmico","term_inflex_m_seco",  "SE/CO Inflex M","MWmed",  "4A2A1E"),
    # ── Política ──────────────────────────────────────────────────────────────
    ("Política","politica_seco",      "Política SE/CO","",       "2A2A4F"),
    ("Política","politica_s",         "Política Sul",  "",       "2A2A4F"),
    ("Política","politica_ne",        "Política NE",   "",       "2A2A4F"),
    ("Política","politica_n",         "Política Norte","",       "2A2A4F"),
]

IS_TEXT = {"arquivo","tipo_doc","mes_pmo","politica_seco","politica_s",
           "politica_ne","politica_n","erro"}


# ══════════════════════════════════════════════════════════════════════════════
# CONSTRUÇÃO DO EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def _sc(ws, row, col, val=None, bold=False, bg=None, fg=CW,
        sz=9, wrap=False, fmt=None, align="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=bold, color=fg, size=sz)
    if bg: c.fill = PatternFill("solid", start_color=bg, end_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fmt: c.number_format = fmt
    return c

def _header_row(ws, campos, row=1):
    """Linha de grupo."""
    prev = None
    for ci,(grp,key,lbl,unit,hcol) in enumerate(campos,1):
        _sc(ws,row,ci, grp if grp!=prev else "",
            bold=True, bg=hcol, fg=CD if hcol==CG else CW, sz=8)
        prev = grp

def _label_row(ws, campos, row=2):
    """Linha de rótulo + unidade."""
    for ci,(_,key,lbl,unit,_) in enumerate(campos,1):
        _sc(ws,row,ci, f"{lbl}\n({unit})" if unit else lbl,
            bold=True, bg=CP, sz=8, wrap=True)

def build_excel(records: List[Dict], output: Path) -> None:
    wb = Workbook()
    thin = Side(style="thin", color="1E3A5F")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Ordenar por data (extraída do conteúdo)
    records.sort(key=lambda r: (
        r.get("ano_pmo") or 0,
        r.get("mes_num") or 0,
        str(r.get("semana_inicio") or ""),
    ))

    # ── Aba 1: Dados PMO (todos os campos) ───────────────────────────────────
    ws = wb.active
    ws.title = "Dados PMO"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"

    _header_row(ws, CAMPOS, 1)
    _label_row(ws, CAMPOS, 2)
    for ci,(_,key,lbl,unit,_) in enumerate(CAMPOS,1):
        ws.column_dimensions[get_column_letter(ci)].width = (
            22 if key.startswith("politica") else
            16 if key in ("arquivo","tipo_doc") else
            10 if "sem" in key or "cmo" in key else 11)
    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 34

    for ri, rec in enumerate(records, 3):
        bg = CP2 if ri%2==1 else CP
        for ci,(_,key,lbl,unit,hcol) in enumerate(CAMPOS,1):
            v = rec.get(key)
            if key in ("semana_inicio","semana_fim") and v: v=str(v)
            txt = key in IS_TEXT
            fmt = None
            if not txt and isinstance(v, float):
                fmt = "#,##0.0" if unit in ("%","%MLT","%EARmax") else "#,##0.00"
            c = _sc(ws,ri,ci,v, bg=bg, sz=9,
                    wrap=key.startswith("politica"),
                    fmt=fmt, align="left" if txt else "center")
            c.border = bdr
            # Alertas
            if "ear_init_rv2_s" in key and isinstance(v,float) and v<35:
                c.fill = PatternFill("solid",start_color=CRE,end_color=CRE)
                c.font = Font(name="Arial",color="FCA5A5",bold=True,size=9)
            elif "cmo" in key and "_s" in key and isinstance(v,float) and v>420:
                c.fill = PatternFill("solid",start_color=COR,end_color=COR)
            elif key=="qualidade_pct" and isinstance(v,float) and v<60:
                c.fill = PatternFill("solid",start_color=CRE,end_color=CRE)
        ws.row_dimensions[ri].height = 15

    # ── Aba 2: Cobertura ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Cobertura")
    ws2.sheet_view.showGridLines = False
    for ci,h in enumerate(["Chave","Rótulo","Preenchidos","Total","Cobertura %"],1):
        _sc(ws2,1,ci,h,bold=True,bg=CG,fg=CD,sz=9)
    for w,col in zip([32,28,13,8,14],"ABCDE"):
        ws2.column_dimensions[col].width=w
    total = len(records)
    for ri,(_,key,lbl,unit,_) in enumerate(CAMPOS,2):
        filled = sum(1 for r in records if r.get(key) not in (None,""))
        pct = filled/total*100 if total else 0
        bg = CP2 if ri%2==0 else CP
        ok = "1A3A1A" if pct>=80 else "3A2A0A" if pct>=40 else "3A0A0A"
        _sc(ws2,ri,1,key,                                bg=bg,sz=8,align="left")
        _sc(ws2,ri,2,f"{lbl} ({unit})" if unit else lbl, bg=bg,sz=8,align="left")
        _sc(ws2,ri,3,filled,                             bg=bg,sz=8)
        _sc(ws2,ri,4,total,                              bg=bg,sz=8)
        _sc(ws2,ri,5,f"{pct:.0f}%",                     bg=ok,bold=True,sz=8)

    # ── Aba 3: Série CMO (histórico semanal) ─────────────────────────────────
    ws3 = wb.create_sheet("Série CMO")
    ws3.sheet_view.showGridLines = False
    h3 = ["Semana início","Mês/Ano","Tipo","SE/CO ant","SE/CO atual",
          "Sul ant","Sul atual","NE ant","NE atual","Norte ant","Norte atual"]
    for ci,h in enumerate(h3,1):
        _sc(ws3,1,ci,h,bold=True,bg=CG,fg=CD,sz=9)
        ws3.column_dimensions[get_column_letter(ci)].width=13
    ws3.column_dimensions["B"].width=12
    ws3.column_dimensions["C"].width=16
    for ri,rec in enumerate(records,2):
        bg = CP2 if ri%2==0 else CP
        mes_ano = f"{rec.get('mes_pmo','?')}/{rec.get('ano_pmo','?')}"
        vals = [str(rec.get("semana_inicio","")) or rec.get("arquivo",""),
                mes_ano, rec.get("tipo_doc","?"),
                rec.get("cmo_ant_seco"),rec.get("cmo_now_seco"),
                rec.get("cmo_ant_s"),   rec.get("cmo_now_s"),
                rec.get("cmo_ant_ne"),  rec.get("cmo_now_ne"),
                rec.get("cmo_ant_n"),   rec.get("cmo_now_n")]
        for ci,v in enumerate(vals,1):
            fmt = "#,##0.00" if ci>3 else None
            _sc(ws3,ri,ci,v,bg=bg,sz=9,fmt=fmt)

    # ── Aba 4: Série ENA ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Série ENA")
    ws4.sheet_view.showGridLines = False
    h4 = ["Semana início","SE/CO sem MW","SE/CO sem %","SE/CO mês %",
          "Sul sem MW","Sul sem %","NE sem MW","NE sem %",
          "Norte sem MW","Norte sem %"]
    for ci,h in enumerate(h4,1):
        _sc(ws4,1,ci,h,bold=True,bg=CG,fg=CD,sz=9)
        ws4.column_dimensions[get_column_letter(ci)].width=13
    for ri,rec in enumerate(records,2):
        bg = CP2 if ri%2==0 else CP
        vals = [str(rec.get("semana_inicio","")) or rec.get("arquivo",""),
                rec.get("ena_prev_sem_seco_mw"),rec.get("ena_prev_sem_seco_mlt"),
                rec.get("ena_prev_mes_seco_mlt"),
                rec.get("ena_prev_sem_s_mw"),   rec.get("ena_prev_sem_s_mlt"),
                rec.get("ena_prev_sem_ne_mw"),  rec.get("ena_prev_sem_ne_mlt"),
                rec.get("ena_prev_sem_n_mw"),   rec.get("ena_prev_sem_n_mlt")]
        for ci,v in enumerate(vals,1):
            fmt = "#,##0.0" if ci in (3,4,6,8,10) else "#,##0" if ci>1 else None
            _sc(ws4,ri,ci,v,bg=bg,sz=9,fmt=fmt)

    # ── Aba 5: Série EAR ─────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Série EAR")
    ws5.sheet_view.showGridLines = False
    h5 = ["Semana início",
          "SE/CO init","SE/CO Inic","SE/CO Fim","SE/CO VE[+1]",
          "Sul init(T6)","Sul fim(T11)",
          "NE init","NE Inic","NE Fim","NE VE[+1]",
          "Norte init","Norte Inic","Norte Fim","Norte VE[+1]"]
    for ci,h in enumerate(h5,1):
        _sc(ws5,1,ci,h,bold=True,bg=CG,fg=CD,sz=9)
        ws5.column_dimensions[get_column_letter(ci)].width=12
    for ri,rec in enumerate(records,2):
        bg = CP2 if ri%2==0 else CP
        vals = [str(rec.get("semana_inicio","")) or rec.get("arquivo",""),
                rec.get("ear_init_rv2_seco"),rec.get("ear_inic_seco"),
                rec.get("ear_fim_mes_seco"), rec.get("ear_ve_abr_seco"),
                rec.get("ear_init_rv2_s"),   rec.get("ear_fim_tab11_s"),
                rec.get("ear_init_rv2_ne"),  rec.get("ear_inic_ne"),
                rec.get("ear_fim_mes_ne"),   rec.get("ear_ve_abr_ne"),
                rec.get("ear_init_rv2_n"),   rec.get("ear_inic_n"),
                rec.get("ear_fim_mes_n"),    rec.get("ear_ve_abr_n")]
        for ci,v in enumerate(vals,1):
            c = _sc(ws5,ri,ci,v,bg=bg,sz=9,fmt="#,##0.0" if ci>1 else None)
            if ci in (5,6) and isinstance(v,float) and v<35:
                c.fill = PatternFill("solid",start_color=CRE,end_color=CRE)
                c.font = Font(name="Arial",color="FCA5A5",bold=True,size=9)

    # ── Aba 6: Diagnóstico de arquivos ────────────────────────────────────────
    ws6 = wb.create_sheet("Diagnóstico")
    ws6.sheet_view.showGridLines = False
    h6 = ["Arquivo","Tipo doc","Semana início","Semana fim","Mês","Ano",
          "Revisão","Páginas","Qualidade %","Erro"]
    for ci,h in enumerate(h6,1):
        _sc(ws6,1,ci,h,bold=True,bg=CG,fg=CD,sz=9)
    ws6.column_dimensions["A"].width=45
    ws6.column_dimensions["B"].width=20
    for col in "CDEFGHIJ": ws6.column_dimensions[col].width=13

    for ri,rec in enumerate(records,2):
        bg = CP2 if ri%2==0 else CP
        q = rec.get("qualidade_pct",0)
        q_bg = bg if q>=80 else "3A2A0A" if q>=40 else "3A0A0A"
        vals = [rec.get("arquivo"),rec.get("tipo_doc"),
                str(rec.get("semana_inicio") or ""),
                str(rec.get("semana_fim") or ""),
                rec.get("mes_pmo"),rec.get("ano_pmo"),
                rec.get("revisao"),rec.get("n_paginas"),
                f"{q}%", rec.get("erro","")]
        for ci,v in enumerate(vals,1):
            b = q_bg if ci==9 else bg
            _sc(ws6,ri,ci,v,bg=b,sz=9,align="left" if ci in (1,2,10) else "center")

    wb.save(str(output))
    print(f"\n✅  Excel salvo: {output}")
    print(f"    Abas: Dados PMO | Cobertura | Série CMO | Série ENA | Série EAR | Diagnóstico")
    print(f"    {len(records)} PMO(s) processados, ordenados por data")



# ══════════════════════════════════════════════════════════════════════════════
# ATUALIZAÇÃO INCREMENTAL
# ══════════════════════════════════════════════════════════════════════════════

def load_existing_semanas(xlsx_path: Path) -> set:
    if not xlsx_path.exists():
        return set()
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(str(xlsx_path), data_only=True)
        ws = wb["Dados PMO"]
        semanas = set()
        for r in range(3, ws.max_row + 1):
            v = ws.cell(r, 3).value
            if v:
                semanas.add(str(v))
        wb.close()
        return semanas
    except Exception as e:
        print(f"  aviso: nao foi possivel ler Excel: {e}")
        return set()


def append_rows_to_excel(new_records: List[Dict], xlsx_path: Path) -> int:
    if not new_records:
        return 0
    import openpyxl as _xl
    wb = _xl.load_workbook(str(xlsx_path))
    thin = Side(style="thin", color="1E3A5F")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb["Dados PMO"]
    for rec in new_records:
        ri = ws.max_row + 1
        bg = CP2 if ri % 2 == 1 else CP
        for ci, (_, key, lbl, unit, hcol) in enumerate(CAMPOS, 1):
            v = rec.get(key)
            if key in ("semana_inicio", "semana_fim") and v:
                v = str(v)
            txt = key in IS_TEXT
            fmt = None
            if not txt and isinstance(v, float):
                fmt = "#,##0.0" if unit in ("%", "%MLT", "%EARmax") else "#,##0.00"
            cell = _sc(ws, ri, ci, v, bg=bg, sz=9,
                       wrap=key.startswith("politica"),
                       fmt=fmt, align="left" if txt else "center")
            cell.border = bdr
            if "ear_init_rv2_s" in key and isinstance(v, float) and v < 35:
                cell.fill = PatternFill("solid", start_color=CRE, end_color=CRE)
                cell.font = Font(name="Arial", color="FCA5A5", bold=True, size=9)
            elif "cmo" in key and "_s" in key and isinstance(v, float) and v > 420:
                cell.fill = PatternFill("solid", start_color=COR, end_color=COR)
        ws.row_dimensions[ri].height = 15

    ws3 = wb["Série CMO"]
    for rec in new_records:
        ri = ws3.max_row + 1
        bg = CP2 if ri % 2 == 0 else CP
        mes_ano = str(rec.get("mes_pmo","?")) + "/" + str(rec.get("ano_pmo","?"))
        vals = [str(rec.get("semana_inicio","")) or rec.get("arquivo",""),
                mes_ano, rec.get("tipo_doc","?"),
                rec.get("cmo_ant_seco"), rec.get("cmo_now_seco"),
                rec.get("cmo_ant_s"),    rec.get("cmo_now_s"),
                rec.get("cmo_ant_ne"),   rec.get("cmo_now_ne"),
                rec.get("cmo_ant_n"),    rec.get("cmo_now_n")]
        for ci, v in enumerate(vals, 1):
            _sc(ws3, ri, ci, v, bg=bg, sz=9, fmt="#,##0.00" if ci > 3 else None)

    ws4 = wb["Série ENA"]
    for rec in new_records:
        ri = ws4.max_row + 1
        bg = CP2 if ri % 2 == 0 else CP
        vals = [str(rec.get("semana_inicio","")) or rec.get("arquivo",""),
                rec.get("ena_prev_sem_seco_mw"),  rec.get("ena_prev_sem_seco_mlt"),
                rec.get("ena_prev_mes_seco_mlt"),
                rec.get("ena_prev_sem_s_mw"),     rec.get("ena_prev_sem_s_mlt"),
                rec.get("ena_prev_sem_ne_mw"),    rec.get("ena_prev_sem_ne_mlt"),
                rec.get("ena_prev_sem_n_mw"),     rec.get("ena_prev_sem_n_mlt")]
        for ci, v in enumerate(vals, 1):
            fmt = "#,##0.0" if ci in (3,4,6,8,10) else "#,##0" if ci > 1 else None
            _sc(ws4, ri, ci, v, bg=bg, sz=9, fmt=fmt)

    ws5 = wb["Série EAR"]
    for rec in new_records:
        ri = ws5.max_row + 1
        bg = CP2 if ri % 2 == 0 else CP
        vals = [str(rec.get("semana_inicio","")) or rec.get("arquivo",""),
                rec.get("ear_init_rv2_seco"), rec.get("ear_inic_seco"),
                rec.get("ear_fim_mes_seco"),  rec.get("ear_ve_abr_seco"),
                rec.get("ear_init_rv2_s"),    rec.get("ear_fim_tab11_s"),
                rec.get("ear_init_rv2_ne"),   rec.get("ear_inic_ne"),
                rec.get("ear_fim_mes_ne"),    rec.get("ear_ve_abr_ne"),
                rec.get("ear_init_rv2_n"),    rec.get("ear_inic_n"),
                rec.get("ear_fim_mes_n"),     rec.get("ear_ve_abr_n")]
        for ci, v in enumerate(vals, 1):
            cell = _sc(ws5, ri, ci, v, bg=bg, sz=9, fmt="#,##0.0" if ci > 1 else None)
            if ci in (5, 6) and isinstance(v, float) and v < 35:
                cell.fill = PatternFill("solid", start_color=CRE, end_color=CRE)
                cell.font = Font(name="Arial", color="FCA5A5", bold=True, size=9)

    ws6 = wb["Diagnóstico"]
    for rec in new_records:
        ri = ws6.max_row + 1
        bg = CP2 if ri % 2 == 0 else CP
        q = rec.get("qualidade_pct", 0)
        vals = [rec.get("arquivo"),      rec.get("tipo_doc"),
                str(rec.get("semana_inicio") or ""),
                str(rec.get("semana_fim") or ""),
                rec.get("mes_pmo"),      rec.get("ano_pmo"),
                rec.get("revisao"),      rec.get("n_paginas"),
                str(q) + "%",            rec.get("erro","")]
        for ci, v in enumerate(vals, 1):
            _sc(ws6, ri, ci, v, bg=bg, sz=9,
                align="left" if ci in (1,2,10) else "center")

    wb.save(str(xlsx_path))
    return len(new_records)


def update_incremental(pmo_dir: Path, xlsx_path: Path,
                       delete_after: bool = False) -> None:
    existing = load_existing_semanas(xlsx_path)
    print("Excel existente: " + str(len(existing)) + " semanas ja registradas")

    pdfs = sorted([
        p for p in pmo_dir.glob("*.pdf")
        if re.search(r"pmo|programa.mensal|sumario.*pmo", p.name, re.I)
    ])
    if not pdfs:
        print("Nenhum PDF encontrado.")
        return

    print("PDFs encontrados: " + str(len(pdfs)))
    new_records = []
    skipped = 0

    for pdf in pdfs:
        rec = parse_pmo(pdf)
        sem = str(rec.get("semana_inicio") or "")
        if sem and sem in existing:
            print("  ja existe: " + sem)
            skipped += 1
            if delete_after:
                pdf.unlink()
        else:
            new_records.append(rec)
            if delete_after and rec.get("qualidade_pct", 0) >= 60:
                pdf.unlink()
                print("  deletado: " + pdf.name)

    print("Ja existiam: " + str(skipped) + " | Novos: " + str(len(new_records)))

    if not new_records:
        print("Excel ja esta atualizado.")
        return

    if not xlsx_path.exists():
        build_excel(new_records, xlsx_path)
        return

    n = append_rows_to_excel(new_records, xlsx_path)
    print(str(n) + " semana(s) adicionada(s) em " + xlsx_path.name)


# ══════════════════════════════════════════════════════════════════════════════
# ENTRYPOINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Parser PMO/ONS -> Excel")
    ap.add_argument("--dir",          default=str(PMO_DIR))
    ap.add_argument("--out",          default=str(OUTPUT_XLS))
    ap.add_argument("--incremental",  action="store_true",
                    help="Adiciona apenas semanas novas ao Excel existente")
    ap.add_argument("--deletar-apos", action="store_true",
                    help="Deleta o PDF apos extracao (usar com --incremental)")
    ap.add_argument("--excluir-2020", action="store_true",
                    help="Excluir PDFs de 2020 (modo completo)")
    args = ap.parse_args()

    pmo_dir  = Path(args.dir)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if args.incremental:
        update_incremental(pmo_dir, out_path,
                           delete_after=getattr(args, "deletar_apos", False))
    else:
        recs = parse_all(pmo_dir)
        if not recs:
            sys.exit(1)
        if args.excluir_2020:
            antes = len(recs)
            recs = [r for r in recs if r.get("ano_pmo") != 2020]
            print("Excluidos " + str(antes-len(recs)) + " PDFs de 2020")
        print(str(len(recs)) + " PMO(s) processados.")
        build_excel(recs, out_path)
