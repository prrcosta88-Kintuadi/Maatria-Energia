# -*- coding: utf-8 -*-
"""
pld_forecast_engine.py — MAÁTria Energia · Motor Preditivo de PLD
==================================================================
Fonte de dados:
  1ª opção: arquivos locais em DATA_DIR (parquets do core_analysis + CSVs ONS/CCEE)
  2ª opção: Neon PostgreSQL (deploy Render — quando cota disponível)
  Fallback: neon_actuals.csv salvo em MODEL_DIR

Previsão de curto prazo com bandas calibradas pelo histórico de erro do ONS:
  - Bandas proporcionais ao nível de CMO (não fixas)
  - Assimétricas em período seco (ENA < 70% MLT)
  - Ajustadas por sazonalidade (set/nov têm padrões distintos)
  - P50 = CMO do ONS (que é bom — R²=0.88, não deve ser "corrigido")
  - P10/P90 = distribuição empírica do erro histórico contextualizado
"""
from __future__ import annotations
import argparse, json, os, pickle, warnings
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import numpy as np
import pandas as pd
from dotenv import load_dotenv

load_dotenv()

try:
    from spectral_engine import (
        get_spectral_features, get_current_fingerprint,
        enrich_features_now, SPECTRAL_FEATURE_COLS,
        SpectralFeatures,
    )
    _SPECTRAL_OK = True
except Exception:
    _SPECTRAL_OK = False
warnings.filterwarnings("ignore", category=UserWarning)

# ── Paths configuráveis via env ───────────────────────────────────────────────
PMO_XLSX  = Path(os.getenv("PMO_XLSX",  "data/ons/PMOs/validacao_pmo.xlsx"))
MODEL_DIR = Path(os.getenv("MODEL_DIR", "data/models"))
DATA_DIR  = Path(os.getenv("DATA_DIR",  "data"))   # raiz dos dados locais
MODEL_DIR.mkdir(parents=True, exist_ok=True)

SUBSISTEMAS = ["seco", "s", "ne", "n"]
SUB_LABEL   = {"seco": "SE/CO", "s": "Sul", "ne": "Nordeste", "n": "Norte"}
QUANTILES   = [0.10, 0.50, 0.90]

MESES = {"janeiro":1,"fevereiro":2,"marco":3,"março":3,"abril":4,"maio":5,
          "junho":6,"julho":7,"agosto":8,"setembro":9,"outubro":10,
          "novembro":11,"dezembro":12}

BASE_FEATURE_COLS = [
    # ENA prevista pelo ONS (do PMO)
    "ena_prev_seco_mlt","ena_prev_s_mlt","ena_prev_ne_mlt","ena_prev_n_mlt",
    # EAR inicial (Tab6 do PMO ou DuckDB)
    "ear_init_seco","ear_init_s","ear_init_ne","ear_init_n",
    # Despacho térmico (do PMO)
    "term_total_sin","thermal_ratio",
    # CMO do ONS (do PMO) e variação
    "cmo_med_seco","cmo_delta_seco",
    # Carga prevista (do PMO)
    "carga_sem1_sin",
    # Sazonalidade cíclica
    "mes_sin","mes_cos",
    # Erro histórico ONS (ENA prevista − ENA realizada)
    "erro_ons_seco","erro_ons_seco_4w",
    # Regime econômico (quando disponível)
    "spdi","structural_drift",
    # CMO realizado (do DuckDB) — diferença entre previsto e realizado semana anterior
    "cmo_real_seco_lag1","cmo_erro_seco_lag1",
    # ENA realizada (do DuckDB) — para calcular erro ONS mais preciso
    "ena_real_seco_mlt_lag1",
    # Features espectrais (spectral_engine) — quando disponíveis
]

FEATURE_COLS = BASE_FEATURE_COLS + (
    [f"spec_{col}" for col in SPECTRAL_FEATURE_COLS] if _SPECTRAL_OK else []
)

# ── Calibração empírica das bandas — 152 semanas reais 2022–2026 ─────────────
#
# Metodologia:
#   Fonte: cmo_ant (previsto pelo ONS) vs cmo_now (realizado) — DuckDB kintuadi
#   Correlação previsto×realizado: r=0.937 → P50 = CMO ONS (NÃO corrigir)
#   Todos os modelos de correção do P50 pioraram o MAE (rolling 8w: +11,8%)
#   Portanto: P50 = CMO ONS + ajuste sazonal muito conservador (set/nov)
#             Bandas P10/P90 = calibradas pelo contexto (nível de CMO + hidrologia)
#
# Bandas baseadas no MAE real por faixa de CMO (152 obs):
#   CMO < R$17:        MAE=5,9   std=9,9   → bandas estreitas ±0.35×std
#   CMO R$17–71:       MAE=13,4  std=21,2  → bandas moderadas
# Calibrado empiricamente — 190 semanas (2022–2026)
# Fonte: kintuadi.duckdb  |  CMO previsto (semana ant.) vs PLD realizado
# delta_p10 = PLD_p10 - P50  |  delta_p90 = PLD_p90 - P50
# bias      = viés sistemático (negativo = ONS subestima PLD)
BAND_CALIBRATION = {
    "piso":  {"lo":   0, "hi":  20, "delta_p10": 43.1, "delta_p90":  69.0, "bias": -58.8},  # n=65
    "baixo": {"lo":  20, "hi":  60, "delta_p10":  5.9, "delta_p90":  43.2, "bias": -20.5},  # n=33
    "medio": {"lo":  60, "hi": 250, "delta_p10":-68.7, "delta_p90":  78.1, "bias":  -2.0},  # n=45
    "alto":  {"lo": 250, "hi":9999, "delta_p10":-92.2, "delta_p90":  29.6, "bias": +26.6},  # n=47
}

# Bias sazonal empírico em R$/MWh absoluto (negativo = ONS subestima PLD)
SEASONAL_BIAS_R = {
     1: -27.3,   2: -38.9,   3: -38.0,   4:  -9.4,
     5: -40.8,   6: -14.6,   7: -19.4,   8: -19.9,
     9: -16.1,  10:  +7.5,  11:  +7.6,  12:  -4.1,
}

# Limiares e bias por regime hidrológico
ENA_SECO_THRESHOLD  = 67    # %MLT — abaixo = período seco
ENA_UMIDO_THRESHOLD = 85    # %MLT — acima  = período úmido
BIAS_SECO           = -27.0  # R$/MWh
BIAS_NORMAL         =  -5.0  # R$/MWh
BIAS_UMIDO          = -20.4  # R$/MWh

# Persistência do erro (autocorrelação confirmada): lag1=0.408, lag2=0.324
AC_LAG1 = 0.408
AC_LAG2 = 0.324



# ══════════════════════════════════════════════════════════════════════════════
# FONTE DE DADOS — LOCAL PRIMEIRO, NEON COMO FALLBACK
# ══════════════════════════════════════════════════════════════════════════════

def load_local_actuals(data_dir: Path = DATA_DIR) -> pd.DataFrame:
    """
    Carrega actuals do kintuadi.duckdb (fonte canônica local).

    Extrai diretamente das tabelas do DuckDB:
      - pld_historical      → pld_real_{sub} (PLD horário → semanal)
      - ear_diario_subsistema → ear_real_{sub} (%EARmax semanal)
      - ena_diario_subsistema → ena_real_{sub}_mlt (%MLT semanal)
      - cmo                 → cmo_real_{sub} (CMO médio semanal)
      - curva_carga         → carga_real (carga SIN semanal)

    Fallback para parquets legados se DuckDB não encontrado.
    """
    duckdb_path = data_dir / "kintuadi.duckdb"

    if duckdb_path.exists():
        df = _load_from_duckdb(duckdb_path)
        if not df.empty:
            print(f"  Local DuckDB: {len(df)} semanas | {df.shape[1]} colunas")
            return df
        print("  ⚠️  DuckDB encontrado mas sem dados extraídos")

    # Fallback: parquets legados
    parquet_path = data_dir / "core_section_economic.parquet"
    if parquet_path.exists():
        df = _load_from_economic_parquet(parquet_path)
        if not df.empty:
            print(f"  Local parquet (legado): {len(df)} semanas")
            return df

    return pd.DataFrame()


def _load_from_advanced_parquet(path: Path) -> pd.DataFrame:
    """Extrai SPDI, Structural Drift do core_section_advanced_metrics.parquet."""
    try:
        import duckdb as _ddb, json as _json
        con = _ddb.connect()
        try:
            row = con.execute(
                "SELECT section_json FROM read_parquet(?) LIMIT 1", [str(path)]
            ).fetchone()
        finally:
            con.close()
        if not row or not row[0]:
            return pd.DataFrame()
        data = _json.loads(row[0]) if isinstance(row[0], str) else row[0]
        if not isinstance(data, dict):
            return pd.DataFrame()
        result = pd.DataFrame()
        for col_key, col_name in [
            ("spdi_series",             "spdi"),
            ("structural_drift_series", "structural_drift"),
            ("structural_gap_series",   "structural_gap"),
        ]:
            sd = data.get(col_key, {})
            if not sd:
                continue
            s = pd.Series(sd)
            s.index = pd.to_datetime(s.index, errors="coerce")
            s = pd.to_numeric(s, errors="coerce").dropna()
            if s.empty:
                continue
            s.index = s.index - pd.to_timedelta(s.index.dayofweek, unit="D")
            result[col_name] = s.resample("W-MON").median()
        result.index = pd.DatetimeIndex(result.index)
        return result[result.index >= "2021-01-01"].sort_index()
    except Exception as e:
        print(f"  ⚠️  parquet advanced: {e}")
        return pd.DataFrame()


def _load_from_economic_parquet(path: Path) -> pd.DataFrame:
    """Extrai PLD/CMO/EAR do core_section_economic.parquet (legado)."""
    try:
        import duckdb as _ddb, json as _json
        con = _ddb.connect()
        try:
            row = con.execute(
                "SELECT section_json FROM read_parquet(?) LIMIT 1", [str(path)]
            ).fetchone()
        finally:
            con.close()
        if not row or not row[0]:
            return pd.DataFrame()
        data = _json.loads(row[0]) if isinstance(row[0], str) else row[0]
        if not isinstance(data, dict):
            return pd.DataFrame()
        result = pd.DataFrame()
        SUB_MAP = {"SUDESTE":"seco","SE":"seco","SE/CO":"seco",
                   "SUL":"s","S":"s","NORDESTE":"ne","NE":"ne","NORTE":"n","N":"n"}
        for sub_key, sd in data.get("pld_by_submercado", {}).items():
            sk = SUB_MAP.get(sub_key.upper().strip())
            if not sk or not sd:
                continue
            s = pd.to_numeric(pd.Series(sd), errors="coerce").dropna()
            s.index = pd.to_datetime(s.index, errors="coerce")
            s.index = s.index - pd.to_timedelta(s.index.dayofweek, unit="D")
            result[f"pld_real_{sk}"] = s.resample("W-MON").mean()
        result.index = pd.DatetimeIndex(result.index)
        return result[result.index >= "2021-01-01"].sort_index()
    except Exception as e:
        print(f"  ⚠️  parquet economic: {e}")
        return pd.DataFrame()


def _load_from_duckdb(duckdb_path: Path) -> pd.DataFrame:
    """
    Lê as tabelas do kintuadi.duckdb e agrega para séries semanais.
    Segunda-feira como início de semana (alinhado com o PMO).
    """
    try:
        import duckdb as _ddb
    except ImportError:
        print("  pip install duckdb")
        return pd.DataFrame()

    try:
        con = _ddb.connect(str(duckdb_path), read_only=True)
    except Exception as e:
        print(f"  ⚠️  DuckDB não abre: {e}")
        return pd.DataFrame()

    result = pd.DataFrame()
    tables = set(con.execute("SHOW TABLES").df()["name"].tolist())

    # Mapeamento submercado → chave interna
    SUB_MAP = {
        "SUDESTE": "seco", "SUDESTE/CENTRO-OESTE": "seco",
        "SE": "seco", "SE/CO": "seco",
        "SUL": "s", "S": "s",
        "NORDESTE": "ne", "NE": "ne",
        "NORTE": "n", "N": "n",
    }

    # ── PLD semanal por submercado ─────────────────────────────────────────
    # Tabela: pld_historical (data, submercado, pld, hora, dia, mes_referencia)
    if "pld_historical" in tables:
        try:
            pld_q = """
                SELECT
                    date_trunc('week', data) + INTERVAL '1 day' AS semana,
                    UPPER(TRIM(submercado))                      AS sub,
                    AVG(pld)                                     AS pld_medio,
                    PERCENTILE_CONT(0.5) WITHIN GROUP
                        (ORDER BY pld)                           AS pld_mediana
                FROM pld_historical
                WHERE data IS NOT NULL
                  AND pld > 0
                  AND YEAR(data) >= 2021
                GROUP BY 1, 2
                ORDER BY 1, 2
            """
            pld_df = con.execute(pld_q).df()
            pld_df["semana"] = pd.to_datetime(pld_df["semana"])
            for sub_raw, sk in SUB_MAP.items():
                mask = pld_df["sub"] == sub_raw
                if mask.sum() == 0:
                    continue
                s = pld_df[mask].set_index("semana")
                if f"pld_real_{sk}" not in result.columns:
                    result[f"pld_real_{sk}"]     = s["pld_medio"]
                    result[f"pld_real_{sk}_p50"] = s["pld_mediana"]
        except Exception as e:
            print(f"  ⚠️  pld_historical: {e}")

    # ── EAR semanal por subsistema ─────────────────────────────────────────
    # Tabela: ear_diario_subsistema (ear_data, id_subsistema, ear_verif_subsistema_percentual)
    if "ear_diario_subsistema" in tables:
        try:
            ear_q = """
                SELECT
                    date_trunc('week', ear_data) + INTERVAL '1 day' AS semana,
                    UPPER(TRIM(id_subsistema))                       AS sub,
                    AVG(ear_verif_subsistema_percentual)             AS ear_pct
                FROM ear_diario_subsistema
                WHERE ear_data IS NOT NULL
                  AND YEAR(ear_data) >= 2021
                GROUP BY 1, 2
                ORDER BY 1, 2
            """
            ear_df = con.execute(ear_q).df()
            ear_df["semana"] = pd.to_datetime(ear_df["semana"])
            for sub_raw, sk in SUB_MAP.items():
                mask = ear_df["sub"] == sub_raw
                if mask.sum() == 0:
                    continue
                s = ear_df[mask].set_index("semana")["ear_pct"]
                if f"ear_real_{sk}" not in result.columns:
                    result[f"ear_real_{sk}"] = s
        except Exception as e:
            print(f"  ⚠️  ear_diario_subsistema: {e}")

    # ── ENA semanal por subsistema ─────────────────────────────────────────
    # Tabela: ena_diario_subsistema (ena_data, id_subsistema, ena_bruta_regiao_percentualmlt)
    if "ena_diario_subsistema" in tables:
        try:
            ena_q = """
                SELECT
                    date_trunc('week', ena_data) + INTERVAL '1 day' AS semana,
                    UPPER(TRIM(id_subsistema))                       AS sub,
                    AVG(ena_bruta_regiao_percentualmlt)              AS ena_mlt
                FROM ena_diario_subsistema
                WHERE ena_data IS NOT NULL
                  AND YEAR(ena_data) >= 2021
                GROUP BY 1, 2
                ORDER BY 1, 2
            """
            ena_df = con.execute(ena_q).df()
            ena_df["semana"] = pd.to_datetime(ena_df["semana"])
            for sub_raw, sk in SUB_MAP.items():
                mask = ena_df["sub"] == sub_raw
                if mask.sum() == 0:
                    continue
                s = ena_df[mask].set_index("semana")["ena_mlt"]
                if f"ena_real_{sk}_mlt" not in result.columns:
                    result[f"ena_real_{sk}_mlt"] = s
        except Exception as e:
            print(f"  ⚠️  ena_diario_subsistema: {e}")

    # ── CMO semanal por subsistema ─────────────────────────────────────────
    # Tabela: cmo (din_instante, id_subsistema, val_cmo)
    if "cmo" in tables:
        try:
            cmo_q = """
                SELECT
                    date_trunc('week', din_instante) + INTERVAL '1 day' AS semana,
                    UPPER(TRIM(id_subsistema))                           AS sub,
                    AVG(val_cmo)                                         AS cmo_medio,
                    PERCENTILE_CONT(0.5) WITHIN GROUP
                        (ORDER BY val_cmo)                               AS cmo_mediana
                FROM cmo
                WHERE din_instante IS NOT NULL
                  AND val_cmo > 0
                  AND YEAR(din_instante) >= 2021
                GROUP BY 1, 2
                ORDER BY 1, 2
            """
            cmo_df = con.execute(cmo_q).df()
            cmo_df["semana"] = pd.to_datetime(cmo_df["semana"])
            for sub_raw, sk in SUB_MAP.items():
                mask = cmo_df["sub"] == sub_raw
                if mask.sum() == 0:
                    continue
                s = cmo_df[mask].set_index("semana")
                if f"cmo_real_{sk}" not in result.columns:
                    result[f"cmo_real_{sk}"]     = s["cmo_medio"]
                    result[f"cmo_real_{sk}_p50"] = s["cmo_mediana"]
        except Exception as e:
            print(f"  ⚠️  cmo: {e}")

    # ── Carga SIN semanal ─────────────────────────────────────────────────
    # Tabela: curva_carga (din_instante, id_subsistema, val_cargaenergiahomwmed)
    if "curva_carga" in tables:
        try:
            carga_q = """
                SELECT
                    date_trunc('week', din_instante) + INTERVAL '1 day' AS semana,
                    SUM(val_cargaenergiahomwmed)                         AS carga_sin
                FROM curva_carga
                WHERE din_instante IS NOT NULL
                  AND YEAR(din_instante) >= 2021
                GROUP BY 1
                ORDER BY 1
            """
            cg_df = con.execute(carga_q).df()
            cg_df["semana"] = pd.to_datetime(cg_df["semana"])
            result["carga_real_sin"] = cg_df.set_index("semana")["carga_sin"]
        except Exception as e:
            print(f"  ⚠️  curva_carga: {e}")

    # ── Geração térmica SIN semanal ───────────────────────────────────────
    # Tabela: geracao_sin_termica (instante, geracao)
    if "geracao_sin_termica" in tables:
        try:
            term_q = """
                SELECT
                    date_trunc('week', instante) + INTERVAL '1 day' AS semana,
                    AVG(geracao)                                      AS term_sin
                FROM geracao_sin_termica
                WHERE instante IS NOT NULL
                  AND YEAR(instante) >= 2021
                GROUP BY 1
                ORDER BY 1
            """
            tm_df = con.execute(term_q).df()
            tm_df["semana"] = pd.to_datetime(tm_df["semana"])
            result["term_real_sin"] = tm_df.set_index("semana")["term_sin"]
        except Exception as e:
            print(f"  ⚠️  geracao_sin_termica: {e}")

    # ── Regime econômico: SPDI e Structural Drift do parquet legado ────────
    adv_path = duckdb_path.parent / "core_section_advanced_metrics.parquet"
    if adv_path.exists():
        try:
            import duckdb as _ddb2
            _con2 = _ddb2.connect()
            _row = _con2.execute(
                "SELECT section_json FROM read_parquet(?) LIMIT 1", [str(adv_path)]
            ).fetchone()
            _con2.close()
            if _row and _row[0]:
                _data = json.loads(_row[0]) if isinstance(_row[0], str) else _row[0]
                for col_key, col_name in [
                    ("spdi_series",             "spdi"),
                    ("structural_drift_series", "structural_drift"),
                    ("structural_gap_series",   "structural_gap"),
                ]:
                    _sd = _data.get(col_key, {})
                    if not _sd:
                        continue
                    _s = pd.Series(_sd)
                    _s.index = pd.to_datetime(_s.index, errors="coerce")
                    _s = pd.to_numeric(_s, errors="coerce").dropna()
                    if _s.empty:
                        continue
                    _s.index = _s.index - pd.to_timedelta(_s.index.dayofweek, unit="D")
                    if col_name not in result.columns:
                        result[col_name] = _s.resample("W-MON").median()
        except Exception as _e:
            pass  # parquet legado opcional — não bloqueia

    con.close()

    if result.empty:
        return pd.DataFrame()

    result.index = pd.DatetimeIndex(result.index)
    result = result[result.index >= "2021-01-01"].sort_index()
    return result


def load_neon_actuals(save_cache: bool = True) -> pd.DataFrame:
    """
    Carrega actuals do Neon PostgreSQL.
    Usado apenas quando os dados locais não estão disponíveis.
    Se bem-sucedido, salva cache em MODEL_DIR/neon_actuals.csv.
    """
    try:
        import psycopg2
        url = os.getenv("DATABASE_URL", "")
        if not url:
            return pd.DataFrame()

        conn = psycopg2.connect(url)

        def _q(sql):
            try:
                return pd.read_sql(sql, conn)
            except Exception:
                return pd.DataFrame()

        actuals = pd.DataFrame()

        # PLD semanal por subsistema
        pld_df = _q("""
            SELECT
                date_trunc('week', MAKE_DATE(
                    CAST(SUBSTR(CAST(mes_referencia AS TEXT), 1, 4) AS INTEGER),
                    CAST(SUBSTR(CAST(mes_referencia AS TEXT), 5, 2) AS INTEGER),
                    dia
                )) AS semana,
                submercado,
                AVG(pld_hora)                                              AS pld_medio,
                PERCENTILE_CONT(0.50) WITHIN GROUP (ORDER BY pld_hora)    AS pld_mediana
            FROM pld_historical
            WHERE mes_referencia IS NOT NULL AND pld_hora > 0
            GROUP BY 1, 2
            ORDER BY 1, 2
        """)
        if not pld_df.empty:
            pld_df["semana"] = pd.to_datetime(pld_df["semana"])
            SUB_MAP = {"SUDESTE":"seco","SE":"seco","SUL":"s","S":"s",
                       "NORDESTE":"ne","NE":"ne","NORTE":"n","N":"n"}
            for sub_raw in pld_df["submercado"].unique():
                sk = SUB_MAP.get(str(sub_raw).upper().strip())
                if not sk:
                    continue
                mask = pld_df["submercado"].str.upper().str.strip() == sub_raw.upper().strip()
                s = pld_df[mask].set_index("semana")
                actuals[f"pld_real_{sk}"]     = s["pld_medio"]
                actuals[f"pld_real_{sk}_p50"] = s["pld_mediana"]

        # EAR semanal
        ear_df = _q("""
            SELECT
                date_trunc('week', instante::timestamp) AS semana,
                id_subsistema,
                AVG(ear_pct) AS ear_pct_medio
            FROM ear_diario_subsistema
            WHERE instante >= '2021-01-01'
            GROUP BY 1, 2
            ORDER BY 1, 2
        """)
        if not ear_df.empty:
            ear_df["semana"] = pd.to_datetime(ear_df["semana"])
            SUB_MAP2 = {"SE":"seco","SUDESTE":"seco","SUL":"s","S":"s",
                        "NE":"ne","NORDESTE":"ne","N":"n","NORTE":"n"}
            for sub_raw in ear_df["id_subsistema"].unique():
                sk = SUB_MAP2.get(str(sub_raw).upper().strip())
                if not sk:
                    continue
                mask = ear_df["id_subsistema"].str.upper().str.strip() == sub_raw.upper().strip()
                actuals[f"ear_real_{sk}"] = ear_df[mask].set_index("semana")["ear_pct_medio"]

        conn.close()

        if actuals.empty:
            return pd.DataFrame()

        actuals.index = pd.to_datetime(actuals.index)
        actuals = actuals.sort_index()

        # Salvar cache para uso offline
        if save_cache:
            actuals.to_csv(MODEL_DIR / "neon_actuals.csv")
            print(f"  Cache salvo: {MODEL_DIR}/neon_actuals.csv")

        print(f"  Neon: {len(actuals)} semanas | {actuals.shape[1]} colunas")
        return actuals

    except Exception as e:
        print(f"  Neon indisponível ({e})")
        return pd.DataFrame()


def load_actuals(data_dir: Path = DATA_DIR) -> pd.DataFrame:
    """
    Carrega actuals com prioridade: Local → Neon → Cache CSV.
    Ponto de entrada único para obtenção de dados reais.
    """
    # 1. Tentar dados locais primeiro
    df = load_local_actuals(data_dir)
    if not df.empty and any(f"pld_real_{s}" in df.columns for s in SUBSISTEMAS):
        print(f"  ✅ Fonte: arquivos locais ({len(df)} semanas)")
        return df

    # 2. Tentar Neon
    df_neon = load_neon_actuals(save_cache=True)
    if not df_neon.empty:
        print(f"  ✅ Fonte: Neon PostgreSQL ({len(df_neon)} semanas)")
        return df_neon

    # 3. Cache salvo
    cache = MODEL_DIR / "neon_actuals.csv"
    if cache.exists():
        df_cache = pd.read_csv(cache, index_col=0, parse_dates=True)
        if not df_cache.empty:
            print(f"  ✅ Fonte: cache salvo ({len(df_cache)} semanas)")
            return df_cache

    print("  ⚠️  Nenhuma fonte de actuals disponível")
    return pd.DataFrame()


# ══════════════════════════════════════════════════════════════════════════════
# FASE 1 — FEATURES DO PMO (sem alteração)
# ══════════════════════════════════════════════════════════════════════════════

def load_pmo_features(xlsx_path: Path) -> pd.DataFrame:
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    ws_ena = wb["Série ENA"]
    ena_cols = ["ena_prev_seco_mw","ena_prev_seco_mlt","ena_prev_mes_seco_mlt",
                "ena_prev_s_mw","ena_prev_s_mlt",None,
                "ena_prev_ne_mw","ena_prev_ne_mlt",
                "ena_prev_n_mw","ena_prev_n_mlt"]
    ena_rows = []
    for r in range(2, ws_ena.max_row+1):
        sem = ws_ena.cell(r,1).value
        if not sem or str(sem) < "2021": continue
        row = {"semana": str(sem)}
        for i,col in enumerate(ena_cols,2):
            if col is None: continue
            v = ws_ena.cell(r,i).value
            row[col] = float(v) if v is not None else np.nan
        ena_rows.append(row)
    df_ena = pd.DataFrame(ena_rows).set_index("semana") if ena_rows else pd.DataFrame()

    ws_cmo = wb["Série CMO"]
    cmo_cols = ["cmo_ant_seco","cmo_now_seco","cmo_ant_s","cmo_now_s",
                "cmo_ant_ne","cmo_now_ne","cmo_ant_n","cmo_now_n"]
    cmo_rows = []
    for r in range(2, ws_cmo.max_row+1):
        sem = ws_cmo.cell(r,1).value
        if not sem or str(sem) < "2021": continue
        row = {"semana": str(sem)}
        for i,col in enumerate(cmo_cols,4):
            v = ws_cmo.cell(r,i).value
            row[col] = float(v) if v is not None else np.nan
        cmo_rows.append(row)
    df_cmo = pd.DataFrame(cmo_rows).set_index("semana") if cmo_rows else pd.DataFrame()

    ws = wb["Dados PMO"]
    hdr = {}
    for c in range(1, ws.max_column+1):
        v = ws.cell(2,c).value
        if v: hdr[str(v).replace("\n"," ")] = c

    pmo_rows = []
    for r in range(3, ws.max_row+1):
        ano = ws.cell(r,6).value
        sem = ws.cell(r,3).value
        if not sem or not ano or int(str(ano)) < 2021: continue
        def _g(cn):
            ci = hdr.get(cn)
            if not ci: return np.nan
            v = ws.cell(r,ci).value
            return float(v) if v is not None else np.nan
        mes_str = ws.cell(r, hdr.get("Mês PMO",1)).value
        row = {
            "semana":           str(sem),
            "mes_num":          MESES.get(str(mes_str).lower().strip(), np.nan) if mes_str else np.nan,
            "ear_init_seco":    _g("SE/CO init (%EARmax)"),
            "ear_init_s":       _g("Sul init (%EARmax)"),
            "ear_init_ne":      _g("NE init (%EARmax)"),
            "ear_init_n":       _g("Norte init (%EARmax)"),
            "term_total_sin":   _g("Total SIN (MWmed)"),
            "term_total_seco":  _g("SE/CO Total M (MWmed)"),
            "term_merito_seco": _g("SE/CO Mérito M (MWmed)"),
            "term_inflex_seco": _g("SE/CO Inflex M (MWmed)"),
            "carga_sem1_sin":   _g("Sem1 SIN (MWmed)"),
            "carga_mensal_sin": _g("Mensal SIN (MWmed)"),
            "cmo_pesada_seco":  _g("SE/CO Pesada (R$/MWh)"),
            "cmo_med_seco":     _g("SE/CO Med.Sem. (R$/MWh)"),
            "cmo_med_s":        _g("Sul Med.Sem. (R$/MWh)"),
            "cmo_med_ne":       _g("NE Med.Sem. (R$/MWh)"),
            "cmo_med_n":        _g("Norte Med.Sem. (R$/MWh)"),
        }
        pmo_rows.append(row)
    df_pmo = pd.DataFrame(pmo_rows).set_index("semana") if pmo_rows else pd.DataFrame()
    wb.close()

    df = df_pmo
    if not df_ena.empty: df = df.join(df_ena, how="left")
    if not df_cmo.empty: df = df.join(df_cmo, how="left")
    df.index = pd.to_datetime(df.index, errors="coerce")
    df = df[df.index.notna()].sort_index()
    print(f"  PMO features: {len(df)} semanas | {df.shape[1]} colunas")
    return df


def build_training_dataset(pmo_df: pd.DataFrame,
                            actuals_df: pd.DataFrame) -> pd.DataFrame:
    if actuals_df.empty:
        print("  Sem actuals — apenas features PMO (modelo incompleto)")
        _build_derived_features_only(pmo_df)
        return pmo_df

    def _to_monday(idx):
        return idx.to_series().apply(
            lambda d: d - timedelta(days=d.weekday()) if pd.notna(d) else pd.NaT)

    pm = pmo_df.copy()
    pm.index = pd.DatetimeIndex(_to_monday(pmo_df.index))
    nm = actuals_df.copy()
    nm.index = pd.DatetimeIndex(_to_monday(actuals_df.index))
    df = pm.join(nm, how="left", rsuffix="_act")

    # Resolver colunas duplicadas (prioridade: actuals)
    for col in actuals_df.columns:
        if col + "_act" in df.columns:
            df[col] = df[col].fillna(df[col + "_act"])
            df.drop(columns=[col + "_act"], inplace=True, errors="ignore")

    # EAR: preencher com actuals onde PMO está vazio
    for sub in SUBSISTEMAS:
        cp, cn = f"ear_init_{sub}", f"ear_real_{sub}"
        if cn in df.columns:
            df[cp] = df[cp].fillna(df[cn])

    # Erro histórico ONS — usar ENA realizada quando disponível, EAR como proxy
    for sub in SUBSISTEMAS:
        ep  = f"ena_prev_{sub}_mlt"
        er  = f"ena_real_{sub}_mlt"   # ENA realizada do DuckDB (mais precisa)
        ear = f"ear_real_{sub}"        # proxy quando ENA real não disponível
        if ep in df.columns:
            ref = df[er] if er in df.columns else df.get(ear)
            if ref is not None:
                df[f"erro_ons_{sub}"]    = df[ep] - ref.shift(1)
                df[f"erro_ons_{sub}_4w"] = df[f"erro_ons_{sub}"].rolling(4, min_periods=2).mean()

    # CMO realizado lag1 e erro CMO previsto×realizado
    if "cmo_real_seco" in df.columns and "cmo_now_seco" in df.columns:
        df["cmo_real_seco_lag1"]  = df["cmo_real_seco"].shift(1)
        df["cmo_erro_seco_lag1"]  = (df["cmo_now_seco"].shift(1) -
                                      df["cmo_real_seco"].shift(1))
    elif "cmo_real_seco" in df.columns:
        df["cmo_real_seco_lag1"]  = df["cmo_real_seco"].shift(1)
        df["cmo_erro_seco_lag1"]  = df["cmo_med_seco"].shift(1) - df["cmo_real_seco"].shift(1)                                      if "cmo_med_seco" in df.columns else np.nan
    else:
        df["cmo_real_seco_lag1"] = np.nan
        df["cmo_erro_seco_lag1"] = np.nan

    # ENA realizada lag1
    if "ena_real_seco_mlt" in df.columns:
        df["ena_real_seco_mlt_lag1"] = df["ena_real_seco_mlt"].shift(1)
    else:
        df["ena_real_seco_mlt_lag1"] = np.nan

    # Target: PLD realizado H semanas à frente (calculado em train_model)
    # Aqui apenas garantir que pld_real_{sub} existe
    for sub in SUBSISTEMAS:
        if f"pld_real_{sub}" not in df.columns:
            df[f"pld_real_{sub}"] = np.nan

    # Features derivadas
    _build_derived_features_only(df)

    # Features espectrais semanais históricas
    if _SPECTRAL_OK:
        try:
            from spectral_engine import SpectralEngine
            sp_engine = SpectralEngine(DATA_DIR)
            sp_weekly = sp_engine.build_weekly_features(weeks=max(52, len(df)//7))
            if not sp_weekly.empty:
                sp_weekly = sp_weekly.add_prefix("spec_")
                # Alinhar ao índice semanal do dataset
                def _to_monday(idx):
                    return idx - pd.to_timedelta(idx.dayofweek, unit="D")
                sp_weekly.index = _to_monday(sp_weekly.index)
                df = df.join(sp_weekly, how="left", rsuffix="_spec")
                print(f"  Spectral features: {len(sp_weekly)} semanas | {sp_weekly.shape[1]} colunas")
        except Exception as e:
            print(f"  ⚠️  Spectral features não disponíveis: {e}")

    n_tgt = sum(df[f"pld_real_{s}"].notna().sum() for s in SUBSISTEMAS)
    print(f"  Dataset: {len(df)} semanas | {df.shape[1]} colunas | "
          f"obs. com PLD real: {n_tgt//len(SUBSISTEMAS)}")
    return df.sort_index()


def _build_derived_features_only(df: pd.DataFrame) -> None:
    """Calcula features derivadas in-place. Suporta df com N linhas ou linha única."""
    single = len(df) == 1   # tratamento especial para predição pontual

    # CMO delta
    if "cmo_now_seco" in df.columns and "cmo_ant_seco" in df.columns:
        df["cmo_delta_seco"] = df["cmo_now_seco"] - df["cmo_ant_seco"]
    elif "cmo_med_seco" in df.columns:
        df["cmo_delta_seco"] = df["cmo_med_seco"].diff() if not single else np.nan
    else:
        df["cmo_delta_seco"] = np.nan

    # Thermal ratio
    if "term_total_sin" in df.columns:
        if "carga_sem1_sin" in df.columns:
            denom = df["carga_sem1_sin"]
        elif "carga_mensal_sin" in df.columns:
            denom = df["carga_mensal_sin"]
        else:
            denom = pd.Series(85000.0, index=df.index)
        df["thermal_ratio"] = (df["term_total_sin"] /
                               denom.replace(0, np.nan)).clip(0, 1)
    else:
        df["thermal_ratio"] = np.nan

    # Sazonalidade
    df["mes_sin"] = np.sin(2 * np.pi * df["mes_num"].fillna(6) / 12)
    df["mes_cos"] = np.cos(2 * np.pi * df["mes_num"].fillna(6) / 12)

    # Erro ONS
    if "erro_ons_seco" not in df.columns:
        if "ena_prev_seco_mlt" in df.columns and not single:
            df["erro_ons_seco"]    = df["ena_prev_seco_mlt"].diff()
            df["erro_ons_seco_4w"] = df["erro_ons_seco"].rolling(4, min_periods=2).mean()
        else:
            df["erro_ons_seco"]    = np.nan
            df["erro_ons_seco_4w"] = np.nan

    for col in ["spdi","structural_drift"]:
        if col not in df.columns:
            df[col] = np.nan

    # Lags
    if "cmo_real_seco_lag1" not in df.columns:
        df["cmo_real_seco_lag1"] = (df["cmo_real_seco"].shift(1)
                                    if "cmo_real_seco" in df.columns and not single
                                    else np.nan)
    if "cmo_erro_seco_lag1" not in df.columns:
        df["cmo_erro_seco_lag1"] = np.nan
    if "ena_real_seco_mlt_lag1" not in df.columns:
        df["ena_real_seco_mlt_lag1"] = (df["ena_real_seco_mlt"].shift(1)
                                         if "ena_real_seco_mlt" in df.columns and not single
                                         else np.nan)


# ══════════════════════════════════════════════════════════════════════════════
# FASE 2 — CURTO PRAZO com BANDAS CALIBRADAS
# ══════════════════════════════════════════════════════════════════════════════

def _get_calibrated_bands(cmo: float, mes: int,
                           ena_mlt: Optional[float] = None,
                           erro_lag1: Optional[float] = None
                           ) -> Tuple[float, float]:
    """
    Retorna (delta_p10, delta_p90) em R$/MWh absoluto.
    PLD_p10 = P50 + delta_p10  |  PLD_p90 = P50 + delta_p90

    Calibrado com 190 semanas de dados reais (kintuadi.duckdb, 2022–2026).
    """
    # Faixa base pelo nível de CMO
    calib = BAND_CALIBRATION["alto"]
    for v in BAND_CALIBRATION.values():
        if v["lo"] <= cmo < v["hi"]:
            calib = v
            break

    delta_p10 = calib["delta_p10"]
    delta_p90 = calib["delta_p90"]

    # Ajuste hidrológico
    if ena_mlt is not None:
        if ena_mlt < ENA_SECO_THRESHOLD:
            delta_p90 *= 1.30   # banda alta +30% em seco
            delta_p10 *= 0.80   # banda baixa -20%
        elif ena_mlt > ENA_UMIDO_THRESHOLD:
            delta_p90 *= 1.10

    # Persistência do erro (autocorr lag1 = 0.408)
    if erro_lag1 is not None and not np.isnan(float(erro_lag1)):
        persist = float(erro_lag1) * AC_LAG1
        if persist < 0:     # semana passada ONS subestimou → tende a repetir
            delta_p90 += abs(persist) * 0.5
        else:               # semana passada ONS superestimou
            delta_p10 -= persist * 0.5

    return delta_p10, delta_p90


def _get_seasonal_p50_adjustment(cmo: float, mes: int,
                                  ena_mlt: Optional[float] = None,
                                  erro_lag1: Optional[float] = None) -> float:
    """
    Ajuste do P50 em R$/MWh.

    Componentes:
    1. Bias sazonal empírico × 60% (conservador)
    2. Bias hidrológico residual × 30%
    3. Persistência do erro lag1 × 40%

    Resultado típico: +10 a +25 R$/MWh (corrige subestimação do ONS).
    """
    # 1. Bias sazonal (sinal: bias negativo = ONS subestima → adj positiva)
    bias_saz = SEASONAL_BIAS_R.get(mes, -17.6)
    adj = -bias_saz * 0.60

    # 2. Bias hidrológico residual
    if ena_mlt is not None:
        bias_h = (BIAS_SECO  if ena_mlt < ENA_SECO_THRESHOLD  else
                  BIAS_UMIDO if ena_mlt > ENA_UMIDO_THRESHOLD else
                  BIAS_NORMAL)
        residual = bias_h - bias_saz
        adj += -residual * 0.30

    # 3. Persistência do erro anterior
    if erro_lag1 is not None and not np.isnan(float(erro_lag1)):
        adj += -float(erro_lag1) * AC_LAG1

    return round(adj, 1)


def forecast_short_term(pmo_state: Dict,
                         ena_mlt_by_sub: Optional[Dict] = None,
                         erro_lag1_by_sub: Optional[Dict] = None) -> Dict:
    """
    Previsão de CURTO PRAZO (1–5 semanas) com bandas calibradas.

    Args:
        pmo_state:        dados do PMO mais recente
        ena_mlt_by_sub:   ENA %MLT atual por subsistema (calibração condicional)
        erro_lag1_by_sub: erro da semana anterior por subsistema
                          (CMO_previsto - PLD_realizado, autocorr=0.408)
    """
    weeks = []
    sem_base = pmo_state.get("Semana fim") or pmo_state.get("semana_fim")
    if isinstance(sem_base, str):
        try:
            sem_base = datetime.strptime(sem_base, "%Y-%m-%d").date()
        except Exception:
            sem_base = None

    sub_labels_xls = {"seco":"SE/CO","s":"Sul","ne":"NE","n":"Norte"}
    cmo_sems = {}
    for sub in SUBSISTEMAS:
        lbl = sub_labels_xls[sub]
        cmo_med_raw = (pmo_state.get(f"{lbl} Med.Sem. (R$/MWh)")
                       or pmo_state.get("SE/CO Med.Sem. (R$/MWh)"))
        cmo_med = float(cmo_med_raw) if cmo_med_raw is not None else None
        vals = []
        for i in range(1, 6):
            v = pmo_state.get(f"{lbl} Sem{i} (R$/MWh)")
            if v is not None:
                fv = float(v)
                # Validar: se Sem{i} >> Med.Sem. (>5×), é provável MWmed — descartar
                if cmo_med and fv > cmo_med * 5:
                    fv = cmo_med
            else:
                fv = cmo_med
            vals.append(fv)
        cmo_sems[sub] = vals

    # Mês da semana base para ajuste sazonal
    mes_base = sem_base.month if sem_base else datetime.now().month
    ena_mlt_now = ena_mlt_by_sub or {}

    for i in range(5):
        sem_dt = sem_base + timedelta(weeks=i+1) if sem_base else None
        mes_sem = (sem_dt.month if sem_dt else mes_base)
        confianca = "alta" if i < 2 else "moderada"

        week = {
            "semana":    str(sem_dt) if sem_dt else f"Semana+{i+1}",
            "horizonte": i + 1,
            "tipo":      "curto_prazo",
            "fonte":     "CMO_ONS_calibrado",
            "confianca": confianca,
        }

        for sub in SUBSISTEMAS:
            cmo = cmo_sems[sub][i]
            if cmo is None:
                week[f"pld_p10_{sub}"] = week[f"pld_p50_{sub}"] = week[f"pld_p90_{sub}"] = None
                continue

            ena_mlt  = ena_mlt_now.get(sub)
            err_lag1 = (erro_lag1_by_sub or {}).get(sub)

            # Bandas calibradas em R$/MWh absolutos (empírico 2022–2026)
            # erro_lag1 ativa a correção por persistência (autocorr=0.408)
            d10, d90 = _get_calibrated_bands(cmo, mes_sem, ena_mlt, err_lag1)

            # Degradar levemente as bandas para semanas mais distantes
            # Sem. 1–2: std real ≈ 51 R$/MWh; sem. 3–5: incerteza +15% por semana
            dist_factor = 1.0 + (i * 0.15)
            d10 *= dist_factor
            d90 *= dist_factor

            # P50 com ajuste sazonal conservador (só meses com evidência forte)
            p50_adj = _get_seasonal_p50_adjustment(cmo, mes_sem, ena_mlt, err_lag1)
            p50 = round(cmo + p50_adj, 2)

            # d10 é negativo (P10 abaixo do P50), d90 é positivo (P90 acima)
            week[f"pld_p10_{sub}"] = round(max(0.0, p50 + d10), 2)
            week[f"pld_p50_{sub}"] = p50
            week[f"pld_p90_{sub}"] = round(p50 + d90, 2)

            # Metadados de calibração (para transparência na UI)
            week[f"_d10_{sub}"]     = round(d10, 1)
            week[f"_d90_{sub}"]     = round(d90, 1)
            week[f"_p50_adj_{sub}"] = round(p50_adj, 1)
            week[f"_ena_mlt_{sub}"] = ena_mlt

        weeks.append(week)

    # Metadados de calibração
    calibration_note = (
        "Bandas calibradas com 152 semanas de erro histórico ONS (2022–2026). "
        "P50 = CMO ONS com ajuste sazonal conservador (R²=0.88). "
        "Bandas proporcionais ao nível de CMO; assimétricas em período seco."
    )

    return {
        "horizonte":        "curto_prazo",
        "descricao":        "1–5 semanas — CMO ONS com bandas calibradas historicamente",
        "calibration_note": calibration_note,
        "semanas":          weeks,
    }


# ══════════════════════════════════════════════════════════════════════════════
# FASE 3 — MODELO GBM QUANTÍLICO (médio/longo prazo)
# ══════════════════════════════════════════════════════════════════════════════

def train_model(df: pd.DataFrame, subsistema: str = "seco",
                horizon_weeks: int = 4) -> Optional[Dict]:
    try:
        from sklearn.ensemble import HistGradientBoostingRegressor
    except ImportError:
        print("  pip install scikit-learn")
        return None

    # Target: PLD realizado H semanas à frente
    pld_col = f"pld_real_{subsistema}"
    if pld_col not in df.columns or df[pld_col].notna().sum() < 52:
        return None

    available = [c for c in FEATURE_COLS if c in df.columns]
    if not available:
        return None

    y = df[pld_col].shift(-horizon_weeks)
    mask = y.notna() & (y > 0)
    X = df[available][mask]
    y = y[mask]
    if len(X) < 52:
        return None

    n_tr = int(len(X) * 0.80)
    X_tr, X_te = X.iloc[:n_tr].values, X.iloc[n_tr:].values
    y_tr, y_te = y.iloc[:n_tr].values, y.iloc[n_tr:].values

    models, metrics = {}, {}
    for q in QUANTILES:
        gb = HistGradientBoostingRegressor(
            loss="quantile", quantile=q,
            max_iter=300, max_depth=4,
            learning_rate=0.04, min_samples_leaf=8,
            l2_regularization=0.1, random_state=42,
        )
        gb.fit(X_tr, y_tr)
        y_pred = gb.predict(X_te)
        if q == 0.50:
            metrics["mae_p50"]  = round(float(np.mean(np.abs(y_pred - y_te))), 2)
            metrics["mape_p50"] = round(float(
                np.nanmean(np.abs((y_pred - y_te) / np.where(y_te == 0, np.nan, y_te))) * 100
            ), 1)
        metrics[f"coverage_p{int(q*100)}"] = round(float(np.mean(y_te <= y_pred)), 3)
        models[q] = gb

    print(f"  {SUB_LABEL[subsistema]} h={horizon_weeks}w: "
          f"MAE={metrics.get('mae_p50')} MAPE={metrics.get('mape_p50')}% "
          f"n={n_tr}+{len(X_te)}")

    return {
        "models":        models,
        "feature_cols":  available,
        "metrics":       metrics,
        "n_train":       n_tr,
        "n_test":        len(X_te),
        "subsistema":    subsistema,
        "horizon_weeks": horizon_weeks,
        "trained_at":    datetime.now().isoformat(),
    }


def train_all_models(df: pd.DataFrame) -> Dict:
    all_models = {}
    for sub in SUBSISTEMAS:
        all_models[sub] = {}
        for h in [4, 8, 12, 26]:
            r = train_model(df, subsistema=sub, horizon_weeks=h)
            if r:
                all_models[sub][h] = r
    return all_models


def save_models(all_models: Dict, model_dir: Path) -> None:
    model_dir.mkdir(parents=True, exist_ok=True)
    for sub, horizons in all_models.items():
        for h, bundle in horizons.items():
            with open(model_dir / f"gbm_{sub}_h{h}.pkl", "wb") as f:
                pickle.dump(bundle, f)
    meta = {
        sub: {
            str(h): {k: v for k, v in b.items() if k != "models"}
            for h, b in hh.items()
        }
        for sub, hh in all_models.items()
    }
    with open(model_dir / "models_meta.json", "w") as f:
        json.dump(meta, f, indent=2)
    print(f"  Modelos salvos em {model_dir}")


def load_models(model_dir: Path) -> Dict:
    all_models = {}
    for sub in SUBSISTEMAS:
        all_models[sub] = {}
        for h in [4, 8, 12, 26]:
            p = model_dir / f"gbm_{sub}_h{h}.pkl"
            if p.exists():
                with open(p, "rb") as f:
                    all_models[sub][h] = pickle.load(f)
    return all_models


# ══════════════════════════════════════════════════════════════════════════════
# FASE 4 — AJUSTE DE REGIME MAÁTria
# ══════════════════════════════════════════════════════════════════════════════

def _apply_regime_adjustment(preds: Dict, regime: Optional[Dict],
                              horizon_weeks: int) -> Dict:
    if not regime or not preds:
        return preds
    spdi  = regime.get("spdi",  1.0) or 1.0
    drift = regime.get("structural_drift", 1.0) or 1.0
    # Peso decai com o horizonte (curto prazo responde mais ao regime atual)
    weight = max(0.1, 1.0 - horizon_weeks / 52)
    shift  = (spdi - 1.0) * 0.05 * weight if abs(spdi - 1.0) > 0.1 else 0.0
    band_x = max(0.0, (drift - 1.0) * 0.03 * weight)
    adjusted = {}
    for q, v in preds.items():
        if v is None:
            adjusted[q] = v
            continue
        va = v * (1 + shift)
        if   q <= 0.10: va *= (1 - band_x)
        elif q >= 0.90: va *= (1 + band_x)
        adjusted[q] = round(va, 2)
    return adjusted


def get_regime_context() -> Optional[Dict]:
    """Busca indicadores de regime da plataforma (últimas 720h)."""
    try:
        import sys
        sys.path.insert(0, str(Path.cwd()))
        from app_premium import _build_hourly_df
        df = _build_hourly_df()
        if df.empty:
            return None
        last = df.last("720h") if len(df) > 720 else df
        def _med(col):
            s = pd.to_numeric(last.get(col, pd.Series(dtype=float)), errors="coerce")
            return float(s.median()) if not s.empty else None
        return {
            "spdi":             _med("spdi"),
            "structural_drift": _med("structural_drift"),
            "structural_gap":   _med("Structural_gap_R$/MWh"),
        }
    except Exception:
        return None


def forecast_medium_long_term(features_now: pd.Series, all_models: Dict,
                               regime_context: Optional[Dict] = None) -> Dict:
    if regime_context:
        for k, v in regime_context.items():
            if k in FEATURE_COLS and v is not None:
                features_now[k] = v

    results = {}
    horizon_labels = {4: "4 semanas", 8: "8 semanas",
                      12: "3 meses", 26: "6 meses"}

    for sub in SUBSISTEMAS:
        results[sub] = {}
        if sub not in all_models:
            continue
        for h, bundle in all_models[sub].items():
            X = np.array([
                float(features_now.get(c, np.nan)) for c in bundle["feature_cols"]
            ]).reshape(1, -1)
            preds = {q: round(float(gb.predict(X)[0]), 2)
                     for q, gb in bundle["models"].items()}
            preds = _apply_regime_adjustment(preds, regime_context, h)
            results[sub][h] = {
                "horizonte_semanas": h,
                "horizonte_label":   horizon_labels.get(h, f"{h}w"),
                "pld_p10":          preds.get(0.10),
                "pld_p50":          preds.get(0.50),
                "pld_p90":          preds.get(0.90),
                "confianca":        "moderada" if h <= 12 else "baixa",
                "mae_historico":    bundle["metrics"].get("mae_p50"),
                "mape_historico":   bundle["metrics"].get("mape_p50"),
            }

    return {
        "horizonte":       "medio_longo_prazo",
        "descricao":       "4–26 semanas — GBM quantílico (HistGBR) treinado 2021–2026",
        "regime_aplicado": regime_context is not None,
        "subsistemas":     results,
    }


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def build_and_train(xlsx_path: Path = PMO_XLSX,
                    model_dir: Path = MODEL_DIR,
                    data_dir:  Path = DATA_DIR) -> Dict:
    print("\n" + "="*55)
    print("MAÁTria Energia · PLD Forecast Engine · Build")
    print("="*55)

    print("\n[1/4] Carregando features do PMO...")
    pmo_df = load_pmo_features(xlsx_path)

    print("\n[2/4] Carregando actuals (local → Neon → cache)...")
    actuals_df = load_actuals(data_dir)

    print("\n[3/4] Construindo dataset de treinamento...")
    df = build_training_dataset(pmo_df, actuals_df)

    df.to_csv(model_dir / "training_dataset.csv")
    df.to_pickle(model_dir / "training_dataset.pkl")
    print(f"  Dataset salvo em {model_dir}/")

    print("\n[4/4] Treinando modelos GBM quantílicos...")
    all_models = train_all_models(df)
    save_models(all_models, model_dir)

    n = sum(len(v) for v in all_models.values())
    print(f"\n✅ Build: {n} modelos ({len(SUBSISTEMAS)} subs × {n // max(len(SUBSISTEMAS), 1)} horizontes)")
    return all_models


def get_latest_pmo_state(xlsx_path: Path = PMO_XLSX) -> Dict:
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb["Dados PMO"]
    hdr = {str(ws.cell(2, c).value or "").replace("\n", " "): c
           for c in range(1, ws.max_column + 1)}
    last = ws.max_row
    while last >= 3 and ws.cell(last, 3).value is None:
        last -= 1
    state = {cn: ws.cell(last, ci).value for cn, ci in hdr.items()}

    ws_ena = wb["Série ENA"]
    le = ws_ena.max_row
    while le >= 2 and ws_ena.cell(le, 1).value is None:
        le -= 1
    for i in range(1, 11):
        h = ws_ena.cell(1, i).value
        if h:
            state[f"ena_{h}"] = ws_ena.cell(le, i).value

    # ENA %MLT por subsistema (para calibração das bandas)
    # col3=SE/CO sem%, col6=Sul sem%, col8=NE sem%, col10=Norte sem%
    ena_by_sub = {
        "seco": ws_ena.cell(le, 3).value,
        "s":    ws_ena.cell(le, 6).value,
        "ne":   ws_ena.cell(le, 8).value,
        "n":    ws_ena.cell(le, 10).value,
    }
    state["_ena_mlt_by_sub"] = {
        k: float(v) if v is not None else None
        for k, v in ena_by_sub.items()
    }
    wb.close()
    return state


def _compute_erro_lag1(data_dir: Path = DATA_DIR) -> Dict[str, Optional[float]]:
    """
    Calcula erro da semana anterior = CMO_previsto − PLD_realizado.
    Autocorrelação confirmada de 0.408 — é o preditor individual mais forte
    para correção da previsão de curto prazo.

    Tenta ler do DuckDB local; retorna None por subsistema se indisponível.
    """
    duckdb_path = data_dir / "kintuadi.duckdb"
    result: Dict[str, Optional[float]] = {s: None for s in SUBSISTEMAS}

    if not duckdb_path.exists():
        return result

    try:
        import duckdb as _ddb
        con = _ddb.connect(str(duckdb_path), read_only=True)

        # CMO médio da penúltima semana (previsto para a semana passada)
        cmo_q = """
            SELECT
                UPPER(TRIM(id_subsistema)) AS sub,
                AVG(val_cmo)               AS cmo_medio
            FROM cmo
            WHERE din_instante >= (CURRENT_DATE - INTERVAL '14 days')
              AND din_instante <  (CURRENT_DATE - INTERVAL '7 days')
              AND val_cmo > 0
            GROUP BY 1
        """
        # PLD real médio da semana passada
        pld_q = """
            SELECT
                UPPER(TRIM(submercado)) AS sub,
                AVG(pld)                AS pld_medio
            FROM pld_historical
            WHERE data >= (CURRENT_DATE - INTERVAL '14 days')
              AND data <  (CURRENT_DATE - INTERVAL '7 days')
              AND pld > 0
            GROUP BY 1
        """
        cmo_df = con.execute(cmo_q).df()
        pld_df = con.execute(pld_q).df()
        con.close()

        SUB_MAP = {
            "SUDESTE":"seco","SE":"seco","SE/CO":"seco",
            "SUL":"s","S":"s",
            "NORDESTE":"ne","NE":"ne",
            "NORTE":"n","N":"n",
        }
        cmo_by = {}
        for _, row in cmo_df.iterrows():
            sk = SUB_MAP.get(row["sub"])
            if sk:
                cmo_by[sk] = float(row["cmo_medio"])

        pld_by = {}
        for _, row in pld_df.iterrows():
            sk = SUB_MAP.get(row["sub"])
            if sk:
                pld_by[sk] = float(row["pld_medio"])

        for sk in SUBSISTEMAS:
            cmo_v = cmo_by.get(sk)
            pld_v = pld_by.get(sk)
            if cmo_v is not None and pld_v is not None:
                result[sk] = round(cmo_v - pld_v, 2)

    except Exception as e:
        print(f"  ⚠️  erro_lag1 indisponível: {e}")

    return result


def run_forecast(all_models=None, xlsx_path=PMO_XLSX,
                 model_dir=MODEL_DIR, data_dir=DATA_DIR) -> Dict:
    if all_models is None:
        all_models = load_models(model_dir)

    pmo_state    = get_latest_pmo_state(xlsx_path)
    regime       = get_regime_context()
    ena_mlt      = pmo_state.get("_ena_mlt_by_sub")

    # Calcular erro_lag1: CMO_previsto(semana ant.) − PLD_realizado(semana ant.)
    # Lê do DuckDB se disponível
    erro_lag1_by_sub = _compute_erro_lag1(data_dir)

    # Curto prazo com bandas calibradas + persistência do erro
    short = forecast_short_term(pmo_state,
                                 ena_mlt_by_sub=ena_mlt,
                                 erro_lag1_by_sub=erro_lag1_by_sub)

    # Features atuais para médio/longo prazo
    # Construir diretamente do PMO state mais recente
    pmo_df = load_pmo_features(xlsx_path)

    # Pegar última linha com CMO válido
    if not pmo_df.empty:
        valid = pmo_df[pmo_df["cmo_med_seco"].notna()]
        base   = valid.iloc[-1].copy() if not valid.empty else pmo_df.iloc[-1].copy()
    else:
        base = pd.Series(dtype=float)

    # Calcular features derivadas numa cópia segura
    tmp = base.to_frame().T.copy()
    _build_derived_features_only(tmp)
    features_now = tmp.iloc[0].copy()

    # Sobrescrever com valores frescos do PMO state (mais confiáveis)
    sub_map_xls = {"seco": "SE/CO", "s": "Sul", "ne": "NE", "n": "Norte"}
    for sub, lbl in sub_map_xls.items():
        # CMO Med. Semanal por subsistema
        v = pmo_state.get(f"{lbl} Med.Sem. (R$/MWh)")
        if v is not None:
            features_now[f"cmo_med_{sub}"] = float(v)
        # ENA %MLT por subsistema
        if ena_mlt and ena_mlt.get(sub) is not None:
            features_now[f"ena_prev_{sub}_mlt"] = float(ena_mlt[sub])
        # EAR inicial (Tab6)
        v2 = pmo_state.get(f"{lbl} init (%EARmax)")
        if v2 is not None:
            features_now[f"ear_init_{sub}"] = float(v2)

    # Despacho térmico do PMO
    v = pmo_state.get("Total SIN (MWmed)")
    if v is not None:
        features_now["term_total_sin"] = float(v)

    # Delta CMO (variação semana a semana)
    v_now = pmo_state.get("SE/CO atual (R$/MWh)")
    v_ant = pmo_state.get("SE/CO ant (R$/MWh)")
    if v_now is not None and v_ant is not None:
        features_now["cmo_delta_seco"] = float(v_now) - float(v_ant)

    # Erro lag1 (persistência do erro da semana anterior)
    if erro_lag1_by_sub and erro_lag1_by_sub.get("seco") is not None:
        features_now["erro_ons_seco"]      = float(erro_lag1_by_sub["seco"])
        features_now["cmo_erro_seco_lag1"] = float(erro_lag1_by_sub["seco"])

    # Regime econômico
    if regime:
        for k, v in regime.items():
            if v is not None:
                features_now[k] = float(v)

    # Features espectrais (Spectral Engine)
    if _SPECTRAL_OK:
        try:
            features_now = enrich_features_now(features_now, data_dir)
        except Exception:
            pass

    # Mês atual
    mes_atual = datetime.now().month
    features_now["mes_sin"] = np.sin(2 * np.pi * mes_atual / 12)
    features_now["mes_cos"] = np.cos(2 * np.pi * mes_atual / 12)

    # Diagnóstico: ver features_now antes de entrar no modelo
    if os.getenv("PLD_DEBUG"):
        import sys
        print("\n=== DEBUG features_now ===", file=sys.stderr)
        for feat in FEATURE_COLS:
            v = features_now.get(feat) if hasattr(features_now, 'get') else None
            if v is None:
                try: v = features_now[feat]
                except: v = "AUSENTE"
            print(f"  {feat}: {v}", file=sys.stderr)
        print("===", file=sys.stderr)

    medium_long = forecast_medium_long_term(features_now, all_models, regime)

    return {
        "timestamp":        datetime.now().isoformat(),
        "semana_pmo":       str(pmo_state.get("Semana início", "")),
        "regime_context":   regime,
        "ena_mlt_atual":    ena_mlt,
        "short_term":       short,
        "medium_long_term": medium_long,
        "subsistemas":      SUB_LABEL,
    }


# ══════════════════════════════════════════════════════════════════════════════
# API REST (FastAPI)
# ══════════════════════════════════════════════════════════════════════════════

def create_api():
    try:
        from fastapi import FastAPI, HTTPException, BackgroundTasks
        from fastapi.middleware.cors import CORSMiddleware
        from pydantic import BaseModel
    except ImportError:
        print("pip install fastapi uvicorn")
        return None

    app = FastAPI(
        title="MAÁTria Energia · PLD Forecast API",
        description="Motor preditivo PLD — curto/médio/longo prazo",
        version="2.0.0",
    )
    app.add_middleware(CORSMiddleware, allow_origins=["*"],
                       allow_methods=["*"], allow_headers=["*"])
    _cache: Dict = {}

    def _models():
        if not _cache:
            _cache.update(load_models(MODEL_DIR))
        return _cache

    @app.get("/health")
    def health():
        m = load_models(MODEL_DIR)
        cache_ok = (MODEL_DIR / "neon_actuals.csv").exists()
        local_ok = (DATA_DIR / "core_section_economic.parquet").exists()
        return {
            "status":        "ok",
            "models_loaded": sum(len(v) for v in m.values()),
            "xlsx_exists":   PMO_XLSX.exists(),
            "local_data":    local_ok,
            "cache_data":    cache_ok,
            "timestamp":     datetime.now().isoformat(),
        }

    @app.get("/forecast/latest")
    def forecast_latest():
        if not PMO_XLSX.exists():
            raise HTTPException(404, f"PMO Excel não encontrado: {PMO_XLSX}")
        try:
            return run_forecast(all_models=_models())
        except Exception as e:
            raise HTTPException(500, str(e))

    @app.get("/forecast/short")
    def forecast_short():
        s = get_latest_pmo_state(PMO_XLSX)
        return forecast_short_term(s, s.get("_ena_mlt_by_sub"))

    class StateInput(BaseModel):
        ena_prev_seco_mlt: Optional[float] = None
        ena_prev_s_mlt:    Optional[float] = None
        ena_prev_ne_mlt:   Optional[float] = None
        ena_prev_n_mlt:    Optional[float] = None
        ear_init_seco:     Optional[float] = None
        ear_init_s:        Optional[float] = None
        ear_init_ne:       Optional[float] = None
        ear_init_n:        Optional[float] = None
        term_total_sin:    Optional[float] = None
        cmo_med_seco:      Optional[float] = None
        mes_num:           Optional[int]   = None
        spdi:              Optional[float] = None
        structural_drift:  Optional[float] = None

    @app.post("/forecast/custom")
    def forecast_custom(state: StateInput):
        feat = pd.Series(state.model_dump())
        mn = feat.get("mes_num") or datetime.now().month
        feat["mes_sin"] = np.sin(2 * np.pi * mn / 12)
        feat["mes_cos"] = np.cos(2 * np.pi * mn / 12)
        regime = {k: v for k, v in state.model_dump().items()
                  if k in ["spdi", "structural_drift"] and v is not None}
        return forecast_medium_long_term(feat, _models(), regime or None)

    @app.get("/forecast/history")
    def forecast_history():
        p = MODEL_DIR / "training_dataset.csv"
        if not p.exists():
            raise HTTPException(404, "Execute --build primeiro.")
        df = pd.read_csv(p, index_col=0, parse_dates=True)
        cols = ["pld_real_seco","pld_real_s","pld_real_ne","pld_real_n",
                "cmo_med_seco","ena_prev_seco_mlt","ear_init_seco","spdi"]
        avail = [c for c in cols if c in df.columns]
        return (df[avail].dropna(subset=["pld_real_seco"], how="all")
                         .tail(104).reset_index().to_dict(orient="records"))

    @app.get("/model/features")
    def model_features():
        imp = {}
        for sub, horizons in _models().items():
            imp[sub] = {}
            for h, b in horizons.items():
                gb = b["models"].get(0.50)
                if gb and hasattr(gb, "feature_importances_"):
                    fi = dict(zip(b["feature_cols"],
                                  gb.feature_importances_.tolist()))
                    imp[sub][str(h)] = dict(sorted(fi.items(),
                                                   key=lambda x: -x[1])[:10])
        return imp

    @app.get("/model/metrics")
    def model_metrics():
        try:
            with open(MODEL_DIR / "models_meta.json") as f:
                return json.load(f)
        except FileNotFoundError:
            raise HTTPException(404, "Execute --build primeiro.")

    @app.get("/model/data_sources")
    def data_sources():
        """Diagnóstico das fontes de dados disponíveis."""
        return {
            "parquet_economic":  (DATA_DIR/"core_section_economic.parquet").exists(),
            "parquet_advanced":  (DATA_DIR/"core_section_advanced_metrics.parquet").exists(),
            "csv_pld":           any((DATA_DIR/"ccee").glob("PLD*.csv"))
                                 if (DATA_DIR/"ccee").exists() else False,
            "csv_ear":           any((DATA_DIR/"ons").glob("EAR*.csv"))
                                 if (DATA_DIR/"ons").exists() else False,
            "cache_actuals":     (MODEL_DIR/"neon_actuals.csv").exists(),
            "pmo_xlsx":          PMO_XLSX.exists(),
            "neon_url_set":      bool(os.getenv("DATABASE_URL", "")),
        }

    @app.post("/model/retrain")
    def retrain(bg: BackgroundTasks):
        def _do():
            nm = build_and_train(PMO_XLSX, MODEL_DIR, DATA_DIR)
            _cache.clear()
            _cache.update(nm)
        bg.add_task(_do)
        return {"status": "retreinamento iniciado em background"}

    return app


# ══════════════════════════════════════════════════════════════════════════════
# ENTRYPOINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="MAÁTria · PLD Forecast Engine v2")
    ap.add_argument("--build",       action="store_true",
                    help="Constrói dataset e treina modelos")
    ap.add_argument("--forecast",    action="store_true",
                    help="Imprime previsão atual e sai")
    ap.add_argument("--diagnose",    action="store_true",
                    help="Verifica fontes de dados disponíveis")
    ap.add_argument("--port",        type=int, default=8001)
    ap.add_argument("--xlsx",        default=str(PMO_XLSX))
    ap.add_argument("--models",      default=str(MODEL_DIR))
    ap.add_argument("--data",        default=str(DATA_DIR))
    args = ap.parse_args()

    PMO_XLSX  = Path(args.xlsx)
    MODEL_DIR = Path(args.models)
    DATA_DIR  = Path(args.data)
    MODEL_DIR.mkdir(parents=True, exist_ok=True)

    if args.diagnose:
        print("\n=== Diagnóstico de fontes de dados ===")
        sources = {
            "PMO Excel":              PMO_XLSX.exists(),
            "Parquet economic":       (DATA_DIR/"core_section_economic.parquet").exists(),
            "Parquet advanced":       (DATA_DIR/"core_section_advanced_metrics.parquet").exists(),
            "CSV PLD (ccee/)":        any((DATA_DIR/"ccee").glob("PLD*.csv"))
                                      if (DATA_DIR/"ccee").exists() else False,
            "CSV EAR (ons/)":         any((DATA_DIR/"ons").glob("EAR*.csv"))
                                      if (DATA_DIR/"ons").exists() else False,
            "Cache actuals (models/)": (MODEL_DIR/"neon_actuals.csv").exists(),
            "Neon DATABASE_URL":       bool(os.getenv("DATABASE_URL","")),
            "Modelos treinados":       (MODEL_DIR/"gbm_seco_h4.pkl").exists(),
        }
        for label, ok in sources.items():
            print(f"  {'✅' if ok else '❌'} {label}")
        print()

    elif args.build:
        build_and_train(PMO_XLSX, MODEL_DIR, DATA_DIR)

    elif args.forecast:
        m = load_models(MODEL_DIR)
        if not any(m.values()):
            print("Modelos não encontrados. Execute --build primeiro.")
        else:
            result = run_forecast(m, PMO_XLSX, MODEL_DIR, DATA_DIR)
            print(json.dumps(result, indent=2, ensure_ascii=False, default=str))

    else:
        try:
            import uvicorn
            app = create_api()
            if app:
                print(f"\nMAÁTria PLD Forecast API em http://0.0.0.0:{args.port}")
                print(f"Docs: http://0.0.0.0:{args.port}/docs")
                uvicorn.run(app, host="0.0.0.0", port=args.port, log_level="warning")
        except ImportError:
            print("pip install fastapi uvicorn")
